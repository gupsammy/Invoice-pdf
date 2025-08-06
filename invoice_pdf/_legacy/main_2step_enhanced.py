import os
import sys
import json
import csv
import time
import logging
import shutil
import asyncio
import argparse
import fitz  # PyMuPDF
import gc  # For garbage collection in chunked processing
import random  # For jitter in retry back-off
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
from dotenv import load_dotenv
from tqdm import tqdm

# Import new config system
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../..'))
from config import Settings
from logging_config import setup_logging
from core.models import ClassificationResult, classification_result_to_dict
from core.rate_limit import CapacityLimiter, retry_with_backoff
import openpyxl
from openpyxl.styles import Font, PatternFill

# Import Google Generative AI
from google import genai
from google.genai import types
import aiohttp

# Import enhanced 2-step prompts v2
from prompts_2step_enhanced_v2 import (
    ENHANCED_CLASSIFICATION_PROMPT_V2,
    ENHANCED_EMPLOYEE_REIMBURSEMENT_EXTRACTION_PROMPT_V2,
    ENHANCED_VENDOR_INVOICE_EXTRACTION_PROMPT_V2
)

# Import streaming CSV writer for Phase 5 optimization
from utilities.streaming_csv import StreamingCSVWriter

# Import SQLite manifest for resumable processing
from utilities.manifest import ProcessingManifest

# Import TUI for interactive monitoring
from utilities.tui import run_tui_monitor

# ===========================
# ENHANCED CONFIGURATION CONSTANTS
# ===========================

# Folder Configuration
INPUT_FOLDER = "input"
OUTPUT_FOLDER = "output"
LOGS_FOLDER = "logs"
JSON_RESPONSES_FOLDER = "json_responses"

# File Configuration
CLASSIFICATION_CSV_FILE = "classification_results.csv"
VENDOR_EXTRACTION_CSV_FILE = "vendor_extraction_results.csv"
EMPLOYEE_EXTRACTION_CSV_FILE = "employee_extraction_results.csv"
COMBINED_CSV_FILE = "combined_enhanced_results.csv"
FAILED_CSV_FILE = "failed_enhanced_files.csv"
LOG_FILE = "invoice_extraction_2step_enhanced.log"

# Load environment variables and initialize settings first
load_dotenv()
settings = Settings.from_env()
API_KEY = settings.gemini_api_key

# Processing Configuration
_MAX_RETRIES = 3
_RETRY_DELAY_BASE_SECONDS = 10
# Phase 1 Optimization: Single quota semaphore for API rate limiting
QUOTA_LIMIT = 10  # Single-key AFC limit (combined for classification + extraction)
PROCESSING_CHUNK_SIZE = settings.processing_chunk_size
# Phase 2 Optimization: Response dump control
SAVE_RESPONSES = settings.debug_responses
# Phase 4 Optimization: PDF file descriptor guard
PDF_FD_SEMAPHORE_LIMIT = 50

# Enhanced 2-Step Configuration
MAX_CLASSIFICATION_PAGES = 7  # Increased from 5 to 7 pages
MAX_EXTRACTION_PAGES = 20     # Only process files with ≤20 pages for extraction (split if >20)
CLASSIFICATION_MODEL = "gemini-2.5-flash"
EXTRACTION_MODEL = "gemini-2.5-pro"

# Concurrency Configuration
MAX_CONCURRENT_CLASSIFY = 5   # Max concurrent classification tasks
MAX_CONCURRENT_EXTRACT = 3    # Max concurrent extraction tasks

# Legacy setup_logging function replaced with new config system
# This is now handled by logging_config.py
pass

# Settings already initialized above

# Module-level PDF file descriptor semaphore (Phase 4 optimization)
# Will be initialized in main()
pdf_fd_semaphore = None

# Module-level processing manifest for resumable operations
# Will be initialized in main()
processing_manifest = None

# Module-level pause/resume events for TUI control
# Will be initialized when TUI is enabled
pause_event = None
shutdown_event = None

def chunker(seq, n):
    """
    Phase 6 optimization: Split sequence into chunks of size n for memory-efficient processing.
    
    Args:
        seq: Sequence to split into chunks
        n: Chunk size
        
    Yields:
        Chunks of the sequence
    """
    for i in range(0, len(seq), n):
        yield seq[i:i+n]

async def safe_get_pdf_page_count(pdf_path: str) -> int:
    """Get PDF page count with file descriptor semaphore guard."""
    global pdf_fd_semaphore
    if pdf_fd_semaphore is None:
        # Fallback to direct call if semaphore not initialized
        return await asyncio.to_thread(get_pdf_page_count, pdf_path)
    
    async with pdf_fd_semaphore:
        return await asyncio.to_thread(get_pdf_page_count, pdf_path)

async def safe_extract_first_n_pages_pdf(pdf_path: str, max_pages: int) -> bytes:
    """Extract PDF pages with file descriptor semaphore guard."""
    global pdf_fd_semaphore
    if pdf_fd_semaphore is None:
        # Fallback to direct call if semaphore not initialized
        return await asyncio.to_thread(extract_first_n_pages_pdf, pdf_path, max_pages)
    
    async with pdf_fd_semaphore:
        return await asyncio.to_thread(extract_first_n_pages_pdf, pdf_path, max_pages)

def get_pdf_page_count(pdf_path: str) -> int:
    """Get the total number of pages in a PDF file."""
    try:
        doc = fitz.open(pdf_path)
        page_count = len(doc)
        doc.close()
        return page_count
    except Exception as e:
        logging.error(f"Error getting page count for {pdf_path}: {str(e)}")
        return 0

def extract_first_n_pages_pdf(pdf_path: str, max_pages: int = MAX_CLASSIFICATION_PAGES) -> bytes:
    """
    Extract the first N pages from a PDF and return as bytes using optimized page slicing.
    
    Args:
        pdf_path: Path to the PDF file
        max_pages: Maximum number of pages to extract (default: 7)
        
    Returns:
        PDF bytes containing only the first N pages
    """
    try:
        # Open the source PDF
        source_doc = fitz.open(pdf_path)
        
        # Determine pages to extract
        pages_to_copy = min(len(source_doc), max_pages)
        
        if pages_to_copy == len(source_doc):
            # If we need all pages, just return the original document as bytes
            pdf_bytes = source_doc.tobytes()
            source_doc.close()
            return pdf_bytes
        
        # Create a new document and copy pages
        new_doc = fitz.open()
        new_doc.insert_pdf(source_doc, from_page=0, to_page=pages_to_copy-1)
        pdf_bytes = new_doc.tobytes()
        new_doc.close()
        source_doc.close()
        return pdf_bytes
        
    except Exception as e:
        logging.error(f"Error extracting first {max_pages} pages from {pdf_path}: {str(e)}")
        raise

def create_preprocessing_failure_result(pdf_path: str, error_message: str) -> ClassificationResult:
    """
    Create a classification result for preprocessing failures.
    Mark as processing_failed without sending to API.
    """
    pdf_path_obj = Path(pdf_path)
    return ClassificationResult(
        file_name=pdf_path_obj.name,
        file_path=pdf_path,
        classification="processing_failed",
        confidence=1.0,
        reasoning=f"Preprocessing failed: {error_message}",
        key_indicators=["PDF preprocessing error"],
        has_employee_codes=False,
        has_vendor_letterhead=False,
        has_invoice_numbers=False,
        has_travel_dates=False,
        appears_financial=False,
        has_amount_calculations=False,
        has_tax_information=False,
        contains_multiple_doc_types=False,
        primary_document_type="preprocessing_error",
        classification_model="preprocessing-error-handler",
        total_pages_in_pdf=0,
        pages_analyzed=0,
        classification_notes=f"File failed PDF preprocessing: {error_message}"
    )

async def classify_document_async(
    pdf_path: str,
    client: "genai.Client",
    capacity_limiter: CapacityLimiter,
    json_responses_folder: str
) -> Optional[Dict[str, Any]]:
    """
    Classify a document using the first 7 pages and enhanced classification prompt.
    
    Args:
        pdf_path: Path to the PDF file
        client: Shared genai client
        capacity_limiter: CapacityLimiter to control concurrent API calls
        json_responses_folder: Folder to save JSON responses
        
    Returns:
        Dictionary containing classification results or None if failed
    """
    pdf_path_obj = Path(pdf_path)
    
    # Check manifest to see if already classified (when resume mode is active)
    global processing_manifest, pause_event, shutdown_event
    if processing_manifest:
        # Use efficient single-file check instead of querying all files
        if processing_manifest.is_file_completed(pdf_path, 'classification'):
            # Already classified, skip
            logging.debug(f"[CLASSIFY] {pdf_path_obj.name} - Already classified, skipping")
            return None
    
    # Check for shutdown request
    if shutdown_event and shutdown_event.is_set():
        logging.info(f"[CLASSIFY] {pdf_path_obj.name} - Shutdown requested, skipping")
        return None
    
    # Wait for pause event (allow processing to continue)
    if pause_event:
        await pause_event.wait()
    
    try:
        # Get total page count - catch preprocessing errors
        try:
            total_pages = await safe_get_pdf_page_count(pdf_path)
            if total_pages == 0:
                error_msg = "Unable to read PDF or empty PDF"
                logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Preprocessing error: {error_msg}")
                result = create_preprocessing_failure_result(pdf_path, error_msg)
                return classification_result_to_dict(result)
        except Exception as e:
            error_msg = f"Failed to get PDF page count: {str(e)}"
            logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Preprocessing error: {error_msg}")
            return create_preprocessing_failure_result(pdf_path, error_msg)
        
        # Extract first 7 pages - catch preprocessing errors (Phase 4: thread-off PDF slicing)
        try:
            pdf_bytes = await safe_extract_first_n_pages_pdf(pdf_path, MAX_CLASSIFICATION_PAGES)
        except Exception as e:
            error_msg = f"Failed to extract PDF pages: {str(e)}"
            logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Preprocessing error: {error_msg}")
            return create_preprocessing_failure_result(pdf_path, error_msg)
        
        contents = [
            types.Part.from_bytes(
                data=pdf_bytes,
                mime_type='application/pdf',
            ),
            f"Classify this document into one of the three categories based on the first {MAX_CLASSIFICATION_PAGES} pages."
        ]
        
        config = types.GenerateContentConfig(system_instruction=ENHANCED_CLASSIFICATION_PROMPT_V2)
        
        # Use the new rate-limited retry logic
        async def classify_operation():
            async with capacity_limiter:
                logging.info(f"[CLASSIFY] {pdf_path_obj.name} - Making API call (Total pages: {total_pages})")
                
                response = await client.aio.models.generate_content(
                    model=CLASSIFICATION_MODEL,
                    contents=contents,
                    config=config
                )
                return response
        
        response = await retry_with_backoff(
            operation=classify_operation,
            max_retries=_MAX_RETRIES,
            base_delay=2.0,
            max_delay=10.0,
            jitter_range=3.0,
            operation_name=f"CLASSIFY {pdf_path_obj.name}"
        )
        
        if response is None:
            return None
        
        # Save response for debugging (controlled by DEBUG_RESPONSES flag)
        success = False  # Will be set based on parsing success
        classification_log_filename = f"{pdf_path_obj.stem}_classification.txt"
        classification_log_path = os.path.join(json_responses_folder, "classification", classification_log_filename)
        
        # Save response only when DEBUG_RESPONSES=1 or on failure
        def save_classification_response():
            os.makedirs(os.path.dirname(classification_log_path), exist_ok=True)
            with open(classification_log_path, 'w', encoding='utf-8') as f:
                f.write(response.text)
        
        # Parse JSON response with robust error handling
        resp_txt = response.text
        json_start = resp_txt.find('{')
        json_end = resp_txt.rfind('}') + 1
        
        if json_start == -1 or json_end == 0:
            if any(err in resp_txt for err in ["502", "Service Unavailable", "server error"]):
                raise RuntimeError("Transient upstream error – will retry")
            logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Unparsable response: {resp_txt[:120]}")
            save_classification_response()  # Save on failure
            return None
        
        try:
            # Cache JSON string once for both parsing attempts
            json_str = resp_txt[json_start:json_end]
            classification_data = json.loads(json_str)
        except json.JSONDecodeError as json_error:
            # Try to fix common JSON malformations
            logging.warning(f"[CLASSIFY] {pdf_path_obj.name} - JSON malformed, attempting repair: {str(json_error)}")
            try:
                import re
                # json_str already cached above
                # Fix common issue: "text" (extra text) -> "text"
                fixed_json = re.sub(r'"([^"]*)" \([^)]*\)', r'"\1"', json_str)
                classification_data = json.loads(fixed_json)
                logging.info(f"[CLASSIFY] {pdf_path_obj.name} - JSON repair successful")
            except json.JSONDecodeError:
                logging.error(f"[CLASSIFY] {pdf_path_obj.name} - JSON repair failed")
                save_classification_response()  # Save on failure
            return None
        
        # Create ClassificationResult from parsed JSON
        classification_result = ClassificationResult(
            file_name=pdf_path_obj.name,
            file_path=str(pdf_path_obj),
            classification=classification_data.get('classification', 'unknown'),
            confidence=classification_data.get('confidence', 0.0),
            reasoning=classification_data.get('reasoning', ''),
            key_indicators=classification_data.get('key_indicators', []),
            classification_model=CLASSIFICATION_MODEL,
            total_pages_in_pdf=total_pages,
            pages_analyzed=min(total_pages, MAX_CLASSIFICATION_PAGES),
            classification_notes=classification_data.get('classification_notes', ''),
            # Flatten document_characteristics or use individual fields
            has_employee_codes=classification_data.get('document_characteristics', {}).get('has_employee_codes', classification_data.get('has_employee_codes', False)),
            has_vendor_letterhead=classification_data.get('document_characteristics', {}).get('has_vendor_letterhead', classification_data.get('has_vendor_letterhead', False)),
            has_invoice_numbers=classification_data.get('document_characteristics', {}).get('has_invoice_numbers', classification_data.get('has_invoice_numbers', False)),
            has_travel_dates=classification_data.get('document_characteristics', {}).get('has_travel_dates', classification_data.get('has_travel_dates', False)),
            appears_financial=classification_data.get('document_characteristics', {}).get('appears_financial', classification_data.get('appears_financial', False)),
            has_amount_calculations=classification_data.get('document_characteristics', {}).get('has_amount_calculations', classification_data.get('has_amount_calculations', False)),
            has_tax_information=classification_data.get('document_characteristics', {}).get('has_tax_information', classification_data.get('has_tax_information', False)),
            contains_multiple_doc_types=classification_data.get('document_characteristics', {}).get('contains_multiple_doc_types', classification_data.get('contains_multiple_doc_types', False)),
            primary_document_type=classification_data.get('document_characteristics', {}).get('primary_document_type', classification_data.get('primary_document_type', 'unknown'))
        )
        
        classification = classification_result.classification
        confidence = classification_result.confidence
        logging.info(f"[CLASSIFY] {pdf_path_obj.name} - Success: {classification} (confidence: {confidence:.2f})")
        
        # Save response only if DEBUG_RESPONSES is enabled (success case)
        if SAVE_RESPONSES:
            save_classification_response()
        
        # Record classification in manifest (when resume mode is active)
        if processing_manifest:
            doc_type = classification_data.get('document_type', '')
            processing_manifest.mark_classified(pdf_path, str(classification), doc_type)
        
        # Convert back to dict for legacy compatibility
        return classification_result_to_dict(classification_result)
        
    except Exception as e:
        error_msg = f"Preprocessing error: {str(e)[:150]}"
        logging.error(f"[CLASSIFY] {pdf_path_obj.name} - {error_msg}")
        
        # Record error in manifest (when resume mode is active)
        if processing_manifest:
            processing_manifest.mark_error(pdf_path, error_msg)
        
        return None

async def extract_document_data_async(
    pdf_path: str,
    document_type: str,
    client: "genai.Client",
    capacity_limiter: CapacityLimiter,
    json_responses_folder: str
) -> Optional[Dict[str, Any]]:
    """
    Extract data from a classified document using enhanced extraction prompts.
    
    Args:
        pdf_path: Path to the PDF file
        document_type: Classification result (employee_t&e or vendor_invoice)
        client: Shared genai client
        capacity_limiter: CapacityLimiter to control concurrent API calls
        json_responses_folder: Folder to save JSON responses
        
    Returns:
        Dictionary containing extracted data or None if failed
    """
    pdf_path_obj = Path(pdf_path)
    
    # Check manifest to see if already extracted (when resume mode is active)
    global processing_manifest, pause_event, shutdown_event
    if processing_manifest:
        # Use efficient single-file check instead of querying all files
        if processing_manifest.is_file_completed(pdf_path, 'extraction'):
            # Already extracted, skip
            logging.debug(f"[EXTRACT] {pdf_path_obj.name} - Already extracted, skipping")
            return None
    
    # Check for shutdown request
    if shutdown_event and shutdown_event.is_set():
        logging.info(f"[EXTRACT] {pdf_path_obj.name} - Shutdown requested, skipping")
        return None
    
    # Wait for pause event (allow processing to continue)
    if pause_event:
        await pause_event.wait()
    
    # Check page count - only process files with ≤20 pages (or first 20 pages if >20)
    total_pages = await safe_get_pdf_page_count(pdf_path)
    process_full_file = total_pages <= MAX_EXTRACTION_PAGES
    
    if total_pages > MAX_EXTRACTION_PAGES:
        logging.info(f"[EXTRACT] {pdf_path_obj.name} - Will use first {MAX_EXTRACTION_PAGES} pages: {total_pages} pages > {MAX_EXTRACTION_PAGES} page limit")
    
    # Select the appropriate extraction prompt
    if document_type == "employee_t&e":
        extraction_prompt = ENHANCED_EMPLOYEE_REIMBURSEMENT_EXTRACTION_PROMPT_V2
    elif document_type == "vendor_invoice":
        extraction_prompt = ENHANCED_VENDOR_INVOICE_EXTRACTION_PROMPT_V2
    else:
        logging.error(f"[EXTRACT] {pdf_path_obj.name} - Invalid document type: {document_type}")
        return None
    
    try:
        # Use full PDF for extraction or first 20 pages if file is oversized
        if process_full_file:
            pdf_bytes = pdf_path_obj.read_bytes()
        else:
            # Extract first 20 pages for oversized files (Phase 4: thread-off PDF slicing)
            pdf_bytes = await safe_extract_first_n_pages_pdf(pdf_path, MAX_EXTRACTION_PAGES)
        
        contents = [
            types.Part.from_bytes(
                data=pdf_bytes,
                mime_type='application/pdf',
            ),
            f"Extract detailed financial data from this {document_type.replace('_', ' ')} document with enhanced accuracy and completeness."
        ]
        
        config = types.GenerateContentConfig(system_instruction=extraction_prompt)
        
        # Use the new rate-limited retry logic
        async def extract_operation():
            async with capacity_limiter:
                logging.info(f"[EXTRACT] {pdf_path_obj.name} - Making API call ({document_type}, {total_pages} pages)")
                
                response = await client.aio.models.generate_content(
                    model=EXTRACTION_MODEL,
                    contents=contents,
                    config=config
                )
                return response
        
        response = await retry_with_backoff(
            operation=extract_operation,
            max_retries=_MAX_RETRIES,
            base_delay=2.0,
            max_delay=10.0,
            jitter_range=3.0,
            operation_name=f"EXTRACT {pdf_path_obj.name} ({document_type})"
        )
        
        if response is None:
            return None
        
        # Save response for debugging in appropriate subfolder
        # Save response for debugging (controlled by DEBUG_RESPONSES flag)
        extraction_log_filename = f"{pdf_path_obj.stem}_extraction_{document_type}.txt"
        extraction_subfolder = os.path.join(json_responses_folder, document_type)
        extraction_log_path = os.path.join(extraction_subfolder, extraction_log_filename)
        
        # Save response only when DEBUG_RESPONSES=1 or on failure
        def save_extraction_response():
            os.makedirs(extraction_subfolder, exist_ok=True)
            with open(extraction_log_path, 'w', encoding='utf-8') as f:
                f.write(response.text)
        
        # Parse JSON response with robust error handling
        resp_txt = response.text
        json_start = resp_txt.find('{')
        json_end = resp_txt.rfind('}') + 1
        
        if json_start == -1 or json_end == 0:
            if any(err in resp_txt for err in ["502", "Service Unavailable", "server error"]):
                raise RuntimeError("Transient upstream error – will retry")
            logging.error(f"[EXTRACT] {pdf_path_obj.name} - Unparsable response: {resp_txt[:120]}")
            save_extraction_response()  # Save on failure
            return None
        
        try:
            # Cache JSON string once for both parsing attempts
            json_str = resp_txt[json_start:json_end]
                        extraction_data = json.loads(json_str)
                    except json.JSONDecodeError as json_error:
                        # Try to fix common JSON malformations
                        logging.warning(f"[EXTRACT] {pdf_path_obj.name} - JSON malformed, attempting repair: {str(json_error)}")
                        try:
                            import re
                            # json_str already cached above
                            # Fix common issue: "text" (extra text) -> "text"
                            fixed_json = re.sub(r'"([^"]*)" \([^)]*\)', r'"\1"', json_str)
                            extraction_data = json.loads(fixed_json)
                            logging.info(f"[EXTRACT] {pdf_path_obj.name} - JSON repair successful")
                        except json.JSONDecodeError:
                            logging.error(f"[EXTRACT] {pdf_path_obj.name} - JSON repair failed")
                            save_extraction_response()  # Save on failure
                            raise json_error
                    
                    # Add metadata
                    extraction_data["file_name"] = pdf_path_obj.name
                    extraction_data["file_path"] = str(pdf_path_obj)
                    extraction_data["extraction_model"] = EXTRACTION_MODEL
                    extraction_data["document_type_processed"] = document_type
                    extraction_data["total_pages_in_pdf"] = total_pages
                    
                    logging.info(f"[EXTRACT] {pdf_path_obj.name} - Success ({document_type})")
                    
                    # Save response only if DEBUG_RESPONSES is enabled (success case)
                    if SAVE_RESPONSES:
                        save_extraction_response()
                    
                    # Record extraction completion in manifest (when resume mode is active)
                    if processing_manifest:
                        processing_manifest.mark_extracted(pdf_path)
                    
                    return extraction_data
                    
                except json.JSONDecodeError as jde:
                    error_msg = f"JSON decode failed: {str(jde)}"
                    logging.error(f"[EXTRACT] {pdf_path_obj.name} - {error_msg}")
                    
                    # Record error in manifest (when resume mode is active)
                    if processing_manifest:
                        processing_manifest.mark_error(pdf_path, error_msg)
                    
                    return None
                except Exception as exc:
                    if attempt < _MAX_RETRIES - 1:
                        # Phase 8: Add jitter to retry back-off
                        base_delay = min(10, 2 ** attempt)
                        jitter = random.uniform(0, 3)
                        delay = base_delay + jitter
                        logging.warning(f"[EXTRACT] {pdf_path_obj.name} - Error: {str(exc)[:100]}. Retrying in {delay:.1f}s...")
                        await asyncio.sleep(delay)
                        continue
                    error_msg = f"Exhausted retries: {str(exc)[:150]}"
                    logging.error(f"[EXTRACT] {pdf_path_obj.name} - {error_msg}")
                    
                    # Record error in manifest (when resume mode is active)
                    if processing_manifest:
                        processing_manifest.mark_error(pdf_path, error_msg)
                    
                    return None
                    
    except Exception as e:
        error_msg = f"Error: {str(e)[:150]}"
        logging.error(f"[EXTRACT] {pdf_path_obj.name} - {error_msg}")
        
        # Record error in manifest (when resume mode is active)
        if processing_manifest:
            processing_manifest.mark_error(pdf_path, error_msg)
        
        return None

async def classify_document_with_metadata(
    pdf_path: str,
    client: "genai.Client",
    semaphore: asyncio.Semaphore,
    json_responses_folder: str
) -> Tuple[str, Optional[Dict[str, Any]]]:
    """
    Wrapper for classify_document_async that returns metadata with result.
    
    Returns:
        Tuple of (pdf_path, classification_result)
    """
    result = await classify_document_async(pdf_path, client, semaphore, json_responses_folder)
    return pdf_path, result

async def extract_document_with_metadata(
    pdf_path: str,
    document_type: str,
    client: "genai.Client",
    semaphore: asyncio.Semaphore,
    json_responses_folder: str
) -> Tuple[Tuple[str, str], Optional[Dict[str, Any]]]:
    """
    Wrapper for extract_document_data_async that returns metadata with result.
    
    Returns:
        Tuple of ((pdf_path, document_type), extraction_result)
    """
    result = await extract_document_data_async(pdf_path, document_type, client, semaphore, json_responses_folder)
    return (pdf_path, document_type), result

async def process_files_batch(
    files: List[Union[str, Tuple[str, str]]],
    client: "genai.Client",
    semaphore: asyncio.Semaphore,
    json_responses_folder: str,
    stage: str = "classification"
) -> Tuple[List[Dict[str, Any]], List[Dict[str, str]]]:
    """
    Unified function to process files through either classification or extraction stage.
    Uses metadata wrapper approach to eliminate task mapping issues.
    
    Args:
        files: List of file paths (for classification) or (file_path, doc_type) tuples (for extraction)
        client: Shared genai client
        semaphore: Semaphore to control concurrent API calls
        json_responses_folder: Folder to save JSON responses
        stage: "classification" or "extraction"
        
    Returns:
        Tuple of (successful_results, failed_files)
    """
    if not files:
        return [], []
    
    # Create appropriate async tasks with metadata wrappers
    if stage == "classification":
        tasks = [
            asyncio.create_task(classify_document_with_metadata(file_path, client, semaphore, json_responses_folder))
            for file_path in files
        ]
        desc = f"Processing {len(files)} files - Classification"
    else:  # extraction
        tasks = [
            asyncio.create_task(extract_document_with_metadata(file_path, doc_type, client, semaphore, json_responses_folder))
            for file_path, doc_type in files
        ]
        desc = f"Processing {len(files)} files - Extraction"
    
    successful_results = []
    failed_files = []
    
    # Use as_completed for streaming progress - metadata approach eliminates task mapping issues
    with tqdm(total=len(tasks), desc=desc, unit="file") as pbar:
        for completed_task in asyncio.as_completed(tasks):
            try:
                # Get metadata and result from wrapper function
                metadata, result = await completed_task
                
                # Extract file info from metadata
                if stage == "classification":
                    file_path = metadata
                    filename = os.path.basename(file_path)
                    doc_type = None
                else:  # extraction
                    file_path, doc_type = metadata
                    filename = os.path.basename(file_path)
                
                if result:
                    # Check for preprocessing failures in classification
                    if stage == "classification" and result.get("preprocessing_failure", False):
                        successful_results.append(result)
                        pbar.set_postfix_str(f"🔧 {filename} (preprocessing error)")
                    else:
                        successful_results.append(result)
                        if stage == "classification":
                            classification = result.get("classification", "unknown")
                            confidence = result.get("confidence", 0)
                            total_pages = result.get("total_pages_in_pdf", "?")
                            pbar.set_postfix_str(f"✅ {filename} ({classification}, {confidence:.2f}, {total_pages}p)")
                        else:  # extraction
                            pbar.set_postfix_str(f"✅ {filename} ({doc_type})")
                else:
                    failed_files.append({
                        "file_name": filename,
                        "file_path": file_path,
                        "failure_stage": stage,
                        "error_message": f"{stage.title()} returned None (JSON decode or API error)",
                        **({"doc_type": doc_type} if stage == "extraction" else {})
                    })
                    pbar.set_postfix_str(f"❌ {filename}")
                    logging.warning(f"[{stage.upper()} FAILED] {filename}")
                    
            except Exception as e:
                # Handle cases where we can't even get metadata (should be rare)
                logging.error(f"[{stage.upper()} ERROR] Task failed completely: {str(e)[:100]}")
                failed_files.append({
                    "file_name": "unknown",
                    "file_path": "unknown", 
                    "failure_stage": stage,
                    "error_message": str(e)[:200],
                    **({"doc_type": "unknown"} if stage == "extraction" else {})
                })
                pbar.set_postfix_str(f"💥 unknown")
            
            pbar.update(1)
    
    return successful_results, failed_files

async def process_with_retries(
    files: List[Union[str, Tuple[str, str]]],
    client: "genai.Client",
    semaphore: asyncio.Semaphore,
    json_responses_folder: str,
    stage: str = "classification",
    max_passes: int = 3
) -> Tuple[List[Dict[str, Any]], List[Dict[str, str]]]:
    """
    Process files with automatic retries using the unified pipeline.
    
    Args:
        files: List of file paths (for classification) or (file_path, doc_type) tuples (for extraction)
        client: Shared genai client
        semaphore: Semaphore to control concurrent API calls
        json_responses_folder: Folder to save JSON responses
        stage: "classification" or "extraction"
        max_passes: Maximum number of retry passes
        
    Returns:
        Tuple of (all_successful_results, final_failed_files)
    """
    all_successful = []
    remaining_files = files
    
    for pass_num in range(max_passes):
        if not remaining_files:
            break
            
        pass_desc = f"Pass {pass_num + 1}/{max_passes}" if pass_num > 0 else "Initial processing"
        if pass_num > 0:
            logging.info(f"🔄 {pass_desc}: Retrying {len(remaining_files)} failed {stage} files...")
        
        successful, failed = await process_files_batch(
            remaining_files, client, semaphore, json_responses_folder, stage
        )
        
        all_successful.extend(successful)
        
        # If no failures, we're done
        if not failed:
            logging.info(f"✅ All {stage} tasks completed successfully!")
            break
            
        # For classification retries, filter out preprocessing failures (they shouldn't be retried)
        if stage == "classification":
            retryable_failed = [
                failure for failure in failed 
                if "preprocessing" not in failure["error_message"].lower()
            ]
            preprocessing_failures = [
                failure for failure in failed 
                if "preprocessing" in failure["error_message"].lower()
            ]
            
            if preprocessing_failures:
                logging.info(f"🔧 Skipping {len(preprocessing_failures)} preprocessing failures (not retryable)")
            
            remaining_files = [failure["file_path"] for failure in retryable_failed]
        else:  # extraction retries
            remaining_files = [
                (failure["file_path"], failure.get("doc_type", "unknown"))
                for failure in failed
            ]
        
        # If no retryable failures, we're done
        if not remaining_files:
            logging.info(f"✅ No retryable {stage} failures remaining")
            break
            
        if pass_num == max_passes - 1:
            logging.warning(f"⚠️  Reached maximum retry passes ({max_passes}) for {stage}")
    
    # Calculate final failures
    final_failed = []
    if stage == "classification":
        # Include both retryable failures that couldn't be recovered and preprocessing failures
        retryable_failed = [
            {"file_name": os.path.basename(file_path), "file_path": file_path, 
             "failure_stage": stage, "error_message": f"Failed after {max_passes} retry passes"}
            for file_path in remaining_files
        ]
        final_failed = retryable_failed
        # Add back preprocessing failures
        if 'preprocessing_failures' in locals():
            final_failed.extend(preprocessing_failures)
    else:
        final_failed = failed
    
    return all_successful, final_failed


def save_vendor_extraction_results_to_csv(extraction_results: List[Dict[str, Any]], output_file: str):
    """Save vendor invoice extraction results to separate CSV."""
    if not extraction_results:
        logging.warning("No vendor extraction results to save.")
        return
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    fieldnames = [
        "File Name", "File Path", "Document Type", "Readable", "Contains Invoices",
        "Multiple Documents", "Orientation Issues", "Data Source", "Issuer", 
        "Consignor", "Consignee", "Vendor Name", "Original Vendor Name", 
        "Invoice Type", "PAN", "Registration Numbers", "Invoice Date", 
        "Document Number", "Invoice Number", "Description", "Basic Amount", 
        "Tax Amount", "Total Amount", "Currency Code", "Original Amount", 
        "Exchange Rate", "Amount Calculated", "Calculation Method", "Is Main Invoice",
        "Total Pages In PDF", "Page Numbers", "Processing Notes"
    ]
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for result in extraction_results:
                doc_status = result.get("document_status", {})
                extracted_data = result.get("extracted_data", [])
                
                if not extracted_data:
                    # Create row with document status only
                    row = {
                        "File Name": result.get("file_name", ""),
                        "File Path": result.get("file_path", ""),
                        "Document Type": result.get("document_type_processed", ""),
                        "Readable": doc_status.get("readable", ""),
                        "Contains Invoices": doc_status.get("contains_invoices", ""),
                        "Multiple Documents": doc_status.get("multiple_documents", ""),
                        "Orientation Issues": doc_status.get("orientation_issues", ""),
                        "Data Source": "",
                        "Issuer": "",
                        "Consignor": "",
                        "Consignee": "",
                        "Vendor Name": "",
                        "Original Vendor Name": "",
                        "Invoice Type": "",
                        "PAN": "",
                        "Registration Numbers": "",
                        "Invoice Date": "",
                        "Document Number": "",
                        "Invoice Number": "",
                        "Description": "",
                        "Basic Amount": "",
                        "Tax Amount": "",
                        "Total Amount": "",
                        "Currency Code": "",
                        "Original Amount": "",
                        "Exchange Rate": "",
                        "Amount Calculated": "",
                        "Calculation Method": "",
                        "Is Main Invoice": "",
                        "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
                        "Page Numbers": "",
                        "Processing Notes": result.get("processing_notes", "")
                    }
                    writer.writerow(row)
                else:
                    # Create row for each extracted data entry
                    for data_entry in extracted_data:
                        reg_numbers = data_entry.get("registration_numbers", [])
                        reg_numbers_str = ", ".join([f"{reg.get('type', '')}:{reg.get('value', '')}" for reg in reg_numbers])
                        
                        page_numbers = data_entry.get("page_numbers", [])
                        page_numbers_str = ", ".join(map(str, page_numbers)) if page_numbers else ""
                        
                        row = {
                            "File Name": result.get("file_name", ""),
                            "File Path": result.get("file_path", ""),
                            "Document Type": result.get("document_type_processed", ""),
                            "Readable": doc_status.get("readable", ""),
                            "Contains Invoices": doc_status.get("contains_invoices", ""),
                            "Multiple Documents": doc_status.get("multiple_documents", ""),
                            "Orientation Issues": doc_status.get("orientation_issues", ""),
                            "Data Source": data_entry.get("data_source", ""),
                            "Issuer": data_entry.get("issuer", ""),
                            "Consignor": data_entry.get("consignor", ""),
                            "Consignee": data_entry.get("consignee", ""),
                            "Vendor Name": data_entry.get("vendor_name", ""),
                            "Original Vendor Name": data_entry.get("original_vendor_name", ""),
                            "Invoice Type": data_entry.get("invoice_type", ""),
                            "PAN": data_entry.get("pan", ""),
                            "Registration Numbers": reg_numbers_str,
                            "Invoice Date": data_entry.get("invoice_date", ""),
                            "Document Number": data_entry.get("document_number", ""),
                            "Invoice Number": data_entry.get("invoice_number", ""),
                            "Description": data_entry.get("description", ""),
                            "Basic Amount": data_entry.get("basic_amount", ""),
                            "Tax Amount": data_entry.get("tax_amount", ""),
                            "Total Amount": data_entry.get("total_amount", ""),
                            "Currency Code": data_entry.get("currency_code", ""),
                            "Original Amount": data_entry.get("original_amount", ""),
                            "Exchange Rate": data_entry.get("exchange_rate", ""),
                            "Amount Calculated": data_entry.get("amount_calculated", ""),
                            "Calculation Method": data_entry.get("calculation_method", ""),
                            "Is Main Invoice": data_entry.get("is_main_invoice", ""),
                            "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
                            "Page Numbers": page_numbers_str,
                            "Processing Notes": result.get("processing_notes", "")
                        }
                        writer.writerow(row)
        
        logging.info(f"Vendor extraction results saved to {output_file}")
    except Exception as e:
        logging.error(f"Error saving vendor extraction results: {str(e)}")

def save_employee_extraction_results_to_csv(extraction_results: List[Dict[str, Any]], output_file: str):
    """Save employee T&E extraction results to separate CSV."""
    if not extraction_results:
        logging.warning("No employee extraction results to save.")
        return
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    fieldnames = [
        "File Name", "File Path", "Document Type", "Total Pages In PDF",
        "Readable", "Contains Invoices", "Multiple Documents", "Orientation Issues", 
        "Data Source", "Employee Name", "Employee Code", "Department", 
        "Invoice Date", "Description", 
        "Basic Amount", "Tax Amount", "Total Amount", "Currency Code", 
        "Original Amount", "Amount Calculated", "Calculation Method",
        "Page Numbers", "Processing Notes"
    ]
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for result in extraction_results:
                doc_status = result.get("document_status", {})
                extracted_data = result.get("extracted_data", [])
                
                if not extracted_data:
                    # Create row with document status only
                    row = {
                        "File Name": result.get("file_name", ""),
                        "File Path": result.get("file_path", ""),
                        "Document Type": result.get("document_type_processed", ""),
                        "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
                        "Readable": doc_status.get("readable", ""),
                        "Contains Invoices": doc_status.get("contains_invoices", ""),
                        "Multiple Documents": doc_status.get("multiple_documents", ""),
                        "Orientation Issues": doc_status.get("orientation_issues", ""),
                        "Data Source": "",
                        "Employee Name": "",
                        "Employee Code": "",
                        "Department": "",
                        "Invoice Date": "",
                        "Description": "",
                        "Basic Amount": "",
                        "Tax Amount": "",
                        "Total Amount": "",
                        "Currency Code": "",
                        "Original Amount": "",
                        "Amount Calculated": "",
                        "Calculation Method": "",
                        "Page Numbers": "",
                        "Processing Notes": result.get("processing_notes", "")
                    }
                    writer.writerow(row)
                else:
                    # Create row for each extracted data entry
                    for data_entry in extracted_data:
                        page_numbers = data_entry.get("page_numbers", [])
                        page_numbers_str = ", ".join(map(str, page_numbers)) if page_numbers else ""
                        
                        row = {
                            "File Name": result.get("file_name", ""),
                            "File Path": result.get("file_path", ""),
                            "Document Type": result.get("document_type_processed", ""),
                            "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
                            "Readable": doc_status.get("readable", ""),
                            "Contains Invoices": doc_status.get("contains_invoices", ""),
                            "Multiple Documents": doc_status.get("multiple_documents", ""),
                            "Orientation Issues": doc_status.get("orientation_issues", ""),
                            "Data Source": data_entry.get("data_source", ""),
                            "Employee Name": data_entry.get("employee_name", ""),
                            "Employee Code": data_entry.get("employee_code", ""),
                            "Department": data_entry.get("department", ""),
                            "Invoice Date": data_entry.get("invoice_date", ""),
                            "Description": data_entry.get("description", ""),
                            "Basic Amount": data_entry.get("basic_amount", ""),
                            "Tax Amount": data_entry.get("tax_amount", ""),
                            "Total Amount": data_entry.get("total_amount", ""),
                            "Currency Code": data_entry.get("currency_code", ""),
                            "Original Amount": data_entry.get("original_amount", ""),
                            "Amount Calculated": data_entry.get("amount_calculated", ""),
                            "Calculation Method": data_entry.get("calculation_method", ""),
                            "Page Numbers": page_numbers_str,
                            "Processing Notes": result.get("processing_notes", "")
                        }
                        writer.writerow(row)
        
        logging.info(f"Employee extraction results saved to {output_file}")
    except Exception as e:
        logging.error(f"Error saving employee extraction results: {str(e)}")


def save_results_to_csv(
    data: List[Dict[str, Any]], 
    output_file: str, 
    field_config: Dict[str, Any],
    data_name: str = "results"
):
    """
    Generic function to save results to CSV.
    
    Args:
        data: List of dictionaries containing the data
        output_file: Output CSV file path
        field_config: Configuration dict with 'fieldnames' and 'row_mapper' function
        data_name: Name for logging (e.g., "classification results", "failed files")
    """
    if not data:
        logging.info(f"No {data_name} to save.")
        return
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    fieldnames = field_config['fieldnames']
    row_mapper = field_config.get('row_mapper')
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            if row_mapper:
                # Check if first row is a dict or list to determine writer type
                first_row = row_mapper(data[0]) if data else {}
                
                if isinstance(first_row, dict):
                    # Use DictWriter for dictionary mappings
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for item in data:
                        rows = row_mapper(item)
                        if isinstance(rows, list):
                            for row in rows:
                                writer.writerow(row)
                        else:
                            writer.writerow(rows)
                else:
                    # Use regular writer for list mappings
                    writer = csv.writer(csvfile)
                    writer.writerow(fieldnames)
                    
                    for item in data:
                        row = row_mapper(item)
                        writer.writerow(row)
            else:
                # Use simple writer for direct field mapping
                writer = csv.writer(csvfile)
                writer.writerow(fieldnames)
                for item in data:
                    row = [item.get(field, "") for field in fieldnames]
                    writer.writerow(row)
        
        logging.info(f"{data_name.title()} saved to {output_file}")
    except Exception as e:
        logging.error(f"Error saving {data_name}: {str(e)}")

# CSV Configuration Objects
def get_classification_csv_config():
    """Configuration for classification results CSV."""
    fieldnames = [
        "File Name", "File Path", "Classification", "Confidence", "Reasoning",
        "Key Indicators", "Has Employee Codes", "Has Vendor Letterhead", 
        "Has Invoice Numbers", "Has Travel Dates", "Appears Financial",
        "Has Amount Calculations", "Has Tax Information", "Contains Multiple Doc Types",
        "Primary Document Type", "Total Pages In PDF", "Pages Analyzed", "Classification Notes"
    ]
    
    def row_mapper(result):
        doc_chars = result.get("document_characteristics", {})
        key_indicators = result.get("key_indicators", [])
        
        return {
            "File Name": result.get("file_name", ""),
            "File Path": result.get("file_path", ""),
            "Classification": result.get("classification", ""),
            "Confidence": result.get("confidence", ""),
            "Reasoning": result.get("reasoning", ""),
            "Key Indicators": ", ".join(key_indicators) if key_indicators else "",
            "Has Employee Codes": doc_chars.get("has_employee_codes", ""),
            "Has Vendor Letterhead": doc_chars.get("has_vendor_letterhead", ""),
            "Has Invoice Numbers": doc_chars.get("has_invoice_numbers", ""),
            "Has Travel Dates": doc_chars.get("has_travel_dates", ""),
            "Appears Financial": doc_chars.get("appears_financial", ""),
            "Has Amount Calculations": doc_chars.get("has_amount_calculations", ""),
            "Has Tax Information": doc_chars.get("has_tax_information", ""),
            "Contains Multiple Doc Types": doc_chars.get("contains_multiple_doc_types", ""),
            "Primary Document Type": doc_chars.get("primary_document_type", ""),
            "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
            "Pages Analyzed": result.get("pages_analyzed", ""),
            "Classification Notes": result.get("classification_notes", "")
        }
    
    return {"fieldnames": fieldnames, "row_mapper": row_mapper}

def get_failed_files_csv_config():
    """Configuration for failed files CSV."""
    fieldnames = ["Failed File Name", "Failed File Path", "Failure Stage", "Error Message"]
    
    def row_mapper(failure):
        return [
            failure["file_name"],
            failure["file_path"],
            failure["failure_stage"],
            failure["error_message"]
        ]
    
    return {"fieldnames": fieldnames, "row_mapper": row_mapper}

def get_vendor_extraction_csv_config():
    """Configuration for vendor extraction results CSV."""
    fieldnames = [
        "file_name", "file_path", "document_type", "readable", "contains_invoices",
        "multiple_documents", "orientation_issues", "data_source", "issuer", 
        "consignor", "consignee", "vendor_name", "original_vendor_name", 
        "invoice_type", "pan", "registration_numbers", "invoice_date", 
        "document_number", "invoice_number", "description", "basic_amount", 
        "tax_amount", "total_amount", "currency_code", "original_amount", 
        "exchange_rate", "amount_calculated", "calculation_method", "is_main_invoice",
        "total_pages_in_pdf", "page_numbers", "processing_notes"
    ]
    return {"fieldnames": fieldnames}

def get_employee_extraction_csv_config():
    """Configuration for employee extraction results CSV - only fields actually extracted."""
    fieldnames = [
        # File metadata
        "file_name", "file_path", "document_type", "total_pages_in_pdf",
        # Document status (from document_status)
        "readable", "contains_invoices", "multiple_documents", "orientation_issues",
        # Core employee data (from extracted_data)
        "data_source", "employee_name", "employee_code", "department", 
        "invoice_date", "description", 
        # Financial data
        "basic_amount", "tax_amount", "total_amount", "currency_code", 
        "original_amount", "amount_calculated", "calculation_method",
        # Processing metadata
        "page_numbers", "processing_notes"
    ]
    return {"fieldnames": fieldnames}

def build_vendor_rows(result: dict):
    """Yield one CSV row per invoice in a VENDOR result."""
    doc = result
    doc_status = doc.get("document_status", {})
    entries = doc.get("extracted_data", []) or [{}]  # at least one row

    for e in entries:
        # Handle registration_numbers list
        reg_numbers = ""
        if e.get("registration_numbers"):
            reg_numbers = ", ".join(
                f"{r.get('type', '')}:{r.get('value', '')}"
                for r in e.get("registration_numbers", [])
            )
        
        # Handle page_numbers list
        page_nums = ""
        if e.get("page_numbers"):
            page_nums = ", ".join(map(str, e.get("page_numbers", [])))
        
        yield {
            "file_name": doc.get("file_name", ""),
            "file_path": doc.get("file_path", ""),
            "document_type": doc.get("document_type_processed", ""),
            "readable": doc_status.get("readable", ""),
            "contains_invoices": doc_status.get("contains_invoices", ""),
            "multiple_documents": doc_status.get("multiple_documents", ""),
            "orientation_issues": doc_status.get("orientation_issues", ""),
            "data_source": e.get("data_source", ""),
            "issuer": e.get("issuer", ""),
            "consignor": e.get("consignor", ""),
            "consignee": e.get("consignee", ""),
            "vendor_name": e.get("vendor_name", ""),
            "original_vendor_name": e.get("original_vendor_name", ""),
            "invoice_type": e.get("invoice_type", ""),
            "pan": e.get("pan", ""),
            "registration_numbers": reg_numbers,
            "invoice_date": e.get("invoice_date", ""),
            "document_number": e.get("document_number", ""),
            "invoice_number": e.get("invoice_number", ""),
            "description": e.get("description", ""),
            "basic_amount": e.get("basic_amount", ""),
            "tax_amount": e.get("tax_amount", ""),
            "total_amount": e.get("total_amount", ""),
            "currency_code": e.get("currency_code", ""),
            "original_amount": e.get("original_amount", ""),
            "exchange_rate": e.get("exchange_rate", ""),
            "amount_calculated": e.get("amount_calculated", ""),
            "calculation_method": e.get("calculation_method", ""),
            "is_main_invoice": e.get("is_main_invoice", ""),
            "total_pages_in_pdf": doc.get("total_pages_in_pdf", ""),
            "page_numbers": page_nums,
            "processing_notes": doc.get("processing_notes", "")
        }

def build_employee_rows(result: dict):
    """Yield one CSV row per expense item in an EMPLOYEE result - only relevant fields."""
    doc = result
    doc_status = doc.get("document_status", {})
    entries = doc.get("extracted_data", []) or [{}]  # at least one row

    for e in entries:
        # Handle page_numbers list
        page_nums = ""
        if e.get("page_numbers"):
            page_nums = ", ".join(map(str, e.get("page_numbers", [])))
        
        yield {
            # File metadata
            "file_name": doc.get("file_name", ""),
            "file_path": doc.get("file_path", ""),
            "document_type": doc.get("document_type_processed", ""),
            "total_pages_in_pdf": doc.get("total_pages_in_pdf", ""),
            # Document status
            "readable": doc_status.get("readable", ""),
            "contains_invoices": doc_status.get("contains_invoices", ""),
            "multiple_documents": doc_status.get("multiple_documents", ""),
            "orientation_issues": doc_status.get("orientation_issues", ""),
            # Core employee data (only fields actually extracted)
            "data_source": e.get("data_source", ""),
            "employee_name": e.get("employee_name", ""),
            "employee_code": e.get("employee_code", ""),
            "department": e.get("department", ""),
            "invoice_date": e.get("invoice_date", ""),
            "description": e.get("description", ""),
            # Financial data
            "basic_amount": e.get("basic_amount", ""),
            "tax_amount": e.get("tax_amount", ""),
            "total_amount": e.get("total_amount", ""),
            "currency_code": e.get("currency_code", ""),
            "original_amount": e.get("original_amount", ""),
            "amount_calculated": e.get("amount_calculated", ""),
            "calculation_method": e.get("calculation_method", ""),
            # Processing metadata
            "page_numbers": page_nums,
            "processing_notes": doc.get("processing_notes", "")
        }

def read_csv_into_rows(file_path: str) -> List[List]:
    """Utility to read a CSV file into a list of rows (including header)."""
    import csv
    rows = []
    if not os.path.exists(file_path):
        return rows
    with open(file_path, 'r', newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    return rows


def create_excel_from_streamed_csvs(output_folder: str, input_folder_name: str):
    """Create an Excel workbook directly from the three streamed CSV files (chunked mode)."""
    import openpyxl
    import csv
    
    workbook = openpyxl.Workbook()
    # Remove default
    workbook.remove(workbook.active)
    
    # Create Not_Processed sheet (filtered classification data)
    classification_csv_path = os.path.join(output_folder, CLASSIFICATION_CSV_FILE)
    if os.path.exists(classification_csv_path):
        ws = workbook.create_sheet("Not_Processed")
        
        # Headers for Not_Processed sheet
        headers = ["File Name", "Classification", "Total Pages In PDF", "Classification Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        # Filter and add only irrelevant and processing_failed rows
        row_idx = 2
        with open(classification_csv_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                classification = row.get('Classification', '')
                if classification in ['irrelevant', 'processing_failed']:
                    filtered_row = [
                        row.get('File Name', ''),
                        row.get('Classification', ''),
                        row.get('Total Pages In PDF', ''),
                        row.get('Classification Notes', '')
                    ]
                    for col, value in enumerate(filtered_row, 1):
                        ws.cell(row=row_idx, column=col, value=value)
                    row_idx += 1
    
    # Create Employee_T&E sheet (remove File Path column)
    employee_csv_path = os.path.join(output_folder, EMPLOYEE_EXTRACTION_CSV_FILE)
    if os.path.exists(employee_csv_path):
        ws = workbook.create_sheet("Employee_T&E")
        
        with open(employee_csv_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = [h for h in reader.fieldnames if h != 'file_path']  # Remove file_path
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            # Add data rows (excluding file_path column)
            row_idx = 2
            f.seek(0)  # Reset file pointer
            reader = csv.DictReader(f)
            for row in reader:
                filtered_row = [row.get(h, '') for h in headers]
                for col, value in enumerate(filtered_row, 1):
                    ws.cell(row=row_idx, column=col, value=value)
                row_idx += 1
    
    # Create Vendor_Invoices sheet (remove File Path column)
    vendor_csv_path = os.path.join(output_folder, VENDOR_EXTRACTION_CSV_FILE)
    if os.path.exists(vendor_csv_path):
        ws = workbook.create_sheet("Vendor_Invoices")
        
        with open(vendor_csv_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = [h for h in reader.fieldnames if h != 'file_path']  # Remove file_path
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="D97230", end_color="D97230", fill_type="solid")
            
            # Add data rows (excluding file_path column)
            row_idx = 2
            f.seek(0)  # Reset file pointer
            reader = csv.DictReader(f)
            for row in reader:
                filtered_row = [row.get(h, '') for h in headers]
                for col, value in enumerate(filtered_row, 1):
                    ws.cell(row=row_idx, column=col, value=value)
                row_idx += 1
    
    if not workbook.worksheets:
        return  # Nothing to save
    excel_filename = f"{input_folder_name}_data.xlsx"
    excel_path = os.path.join(output_folder, excel_filename)
    workbook.save(excel_path)


def create_worksheet(workbook, sheet_name: str, headers: List[str], data_rows: List[List], header_color: str = "366092"):
    """
    Helper function to create a worksheet with headers and data.
    
    Args:
        workbook: openpyxl workbook object
        sheet_name: Name of the worksheet
        headers: List of header strings
        data_rows: List of lists, each containing row data
        header_color: Hex color for header background (default: blue)
    """
    ws = workbook.create_sheet(sheet_name)
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    
    # Add data rows
    for row_idx, row_data in enumerate(data_rows, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    return ws

def create_excel_from_csvs(classification_csv: str, employee_csv: str, vendor_csv: str, output_folder: str, input_folder_name: str):
    """Create a simple Excel file by loading the three CSVs into separate worksheets.
    This is used in chunked streaming mode where we don't keep all data in memory."""
    import csv
    excel_filename = f"{input_folder_name}_data.xlsx"
    excel_path = os.path.join(output_folder, excel_filename)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(csv_path: str, sheet_name: str, header_color: str):
        if not os.path.exists(csv_path):
            return
        ws = wb.create_sheet(sheet_name)
        with open(csv_path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            for r_idx, row in enumerate(reader, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")

    add_sheet(classification_csv, "Classification", "366092")
    add_sheet(employee_csv, "Employee_T&E", "70AD47")
    add_sheet(vendor_csv, "Vendor_Invoices", "D97230")
    wb.save(excel_path)
    logging.info(f"Excel (CSV-based) report saved to {excel_path}")


def create_excel_report(
    classification_results: List[Dict[str, Any]], 
    vendor_results: List[Dict[str, Any]], 
    employee_results: List[Dict[str, Any]], 
    output_folder: str, 
    input_folder_name: str
):
    """Create Excel file with 3 worksheets: Classification (problems only), Employee T&E, and Vendor Invoices."""
    excel_filename = f"{input_folder_name}_data.xlsx"
    excel_path = os.path.join(output_folder, excel_filename)
    
    try:
        # Create workbook and worksheets
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create Not_Processed worksheet - only include irrelevant/problem files
        problem_classifications = [result for result in classification_results 
                                 if result.get("classification") in ["irrelevant", "processing_failed"]]
        
        if problem_classifications:
            headers = ["File Name", "Classification", "Total Pages In PDF", "Classification Notes"]
            data_rows = [
                [
                    result.get("file_name", ""),
                    result.get("classification", ""),
                    result.get("total_pages_in_pdf", ""),
                    result.get("classification_notes", "")
                ]
                for result in problem_classifications
            ]
            create_worksheet(workbook, "Not_Processed", headers, data_rows, "FF6B6B")
        
        # Create Employee T&E worksheet
        if employee_results:
            ws_employee = workbook.create_sheet("Employee_T&E")
            
            # Headers for employee T&E (only include specified headers)
            headers = [
                "File Name", "Document Type", "Readable", "Orientation Issues", 
                "Data Source", "Employee Name", "Employee Code", "Department", 
                "Invoice Date", "Description", "Basic Amount", "Tax Amount", 
                "Total Amount", "Currency Code", "Original Amount", "Amount Calculated", 
                "Calculation Method", "Page Numbers", "Processing Notes"
            ]
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws_employee.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            # Add data
            row_idx = 2
            for result in employee_results:
                doc_status = result.get("document_status", {})
                extracted_data = result.get("extracted_data", [])
                
                if not extracted_data:
                    # Create row with document status only
                    row_data = [
                        result.get("file_name", ""),
                        result.get("document_type_processed", ""),
                        doc_status.get("readable", ""),
                        doc_status.get("orientation_issues", ""),
                        "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                        result.get("processing_notes", "")
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws_employee.cell(row=row_idx, column=col, value=value)
                    row_idx += 1
                else:
                    # Create row for each extracted data entry
                    for data_entry in extracted_data:
                        reg_numbers = data_entry.get("registration_numbers", [])
                        reg_numbers_str = ", ".join([f"{reg.get('type', '')}:{reg.get('value', '')}" for reg in reg_numbers])
                        
                        page_numbers = data_entry.get("page_numbers", [])
                        page_numbers_str = ", ".join(map(str, page_numbers)) if page_numbers else ""
                        
                        row_data = [
                            result.get("file_name", ""),
                            result.get("document_type_processed", ""),
                            doc_status.get("readable", ""),
                            doc_status.get("orientation_issues", ""),
                            data_entry.get("data_source", ""),
                            data_entry.get("employee_name", ""),
                            data_entry.get("employee_code", ""),
                            data_entry.get("department", ""),
                            data_entry.get("invoice_date", ""),
                            data_entry.get("description", ""),
                            data_entry.get("basic_amount", ""),
                            data_entry.get("tax_amount", ""),
                            data_entry.get("total_amount", ""),
                            data_entry.get("currency_code", ""),
                            data_entry.get("original_amount", ""),
                            data_entry.get("amount_calculated", ""),
                            data_entry.get("calculation_method", ""),
                            page_numbers_str,
                            result.get("processing_notes", "")
                        ]
                        
                        for col, value in enumerate(row_data, 1):
                            ws_employee.cell(row=row_idx, column=col, value=value)
                        row_idx += 1
        
        # Create Vendor Invoices worksheet
        if vendor_results:
            ws_vendor = workbook.create_sheet("Vendor_Invoices")
            
            # Headers for vendor invoices (removed File Path)
            headers = [
                "File Name", "Document Type", "Readable", "Contains Invoices",
                "Multiple Documents", "Orientation Issues", "Data Source", "Issuer", 
                "Consignor", "Consignee", "Vendor Name", "Original Vendor Name", 
                "Invoice Type", "PAN", "Registration Numbers", "Invoice Date", 
                "Document Number", "Invoice Number", "Description", "Basic Amount", 
                "Tax Amount", "Total Amount", "Currency Code", "Original Amount", 
                "Exchange Rate", "Amount Calculated", "Calculation Method", "Is Main Invoice",
                "Total Pages In PDF", "Page Numbers", "Processing Notes"
            ]
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws_vendor.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D97230", end_color="D97230", fill_type="solid")
            
            # Add data
            row_idx = 2
            for result in vendor_results:
                doc_status = result.get("document_status", {})
                extracted_data = result.get("extracted_data", [])
                
                if not extracted_data:
                    # Create row with document status only
                    row_data = [
                        result.get("file_name", ""),
                        result.get("document_type_processed", ""),
                        doc_status.get("readable", ""),
                        doc_status.get("contains_invoices", ""),
                        doc_status.get("multiple_documents", ""),
                        doc_status.get("orientation_issues", ""),
                        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                        result.get("total_pages_in_pdf", ""),
                        "",
                        result.get("processing_notes", "")
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws_vendor.cell(row=row_idx, column=col, value=value)
                    row_idx += 1
                else:
                    # Create row for each extracted data entry
                    for data_entry in extracted_data:
                        reg_numbers = data_entry.get("registration_numbers", [])
                        reg_numbers_str = ", ".join([f"{reg.get('type', '')}:{reg.get('value', '')}" for reg in reg_numbers])
                        
                        page_numbers = data_entry.get("page_numbers", [])
                        page_numbers_str = ", ".join(map(str, page_numbers)) if page_numbers else ""
                        
                        row_data = [
                            result.get("file_name", ""),
                            result.get("document_type_processed", ""),
                            doc_status.get("readable", ""),
                            doc_status.get("contains_invoices", ""),
                            doc_status.get("multiple_documents", ""),
                            doc_status.get("orientation_issues", ""),
                            data_entry.get("data_source", ""),
                            data_entry.get("issuer", ""),
                            data_entry.get("consignor", ""),
                            data_entry.get("consignee", ""),
                            data_entry.get("vendor_name", ""),
                            data_entry.get("original_vendor_name", ""),
                            data_entry.get("invoice_type", ""),
                            data_entry.get("pan", ""),
                            reg_numbers_str,
                            data_entry.get("invoice_date", ""),
                            data_entry.get("document_number", ""),
                            data_entry.get("invoice_number", ""),
                            data_entry.get("description", ""),
                            data_entry.get("basic_amount", ""),
                            data_entry.get("tax_amount", ""),
                            data_entry.get("total_amount", ""),
                            data_entry.get("currency_code", ""),
                            data_entry.get("original_amount", ""),
                            data_entry.get("exchange_rate", ""),
                            data_entry.get("amount_calculated", ""),
                            data_entry.get("calculation_method", ""),
                            data_entry.get("is_main_invoice", ""),
                            result.get("total_pages_in_pdf", ""),
                            page_numbers_str,
                            result.get("processing_notes", "")
                        ]
                        
                        for col, value in enumerate(row_data, 1):
                            ws_vendor.cell(row=row_idx, column=col, value=value)
                        row_idx += 1
        
        # Ensure at least one worksheet exists
        if not workbook.worksheets:
            workbook.create_sheet("No_Data")
            workbook.active["A1"] = "No data available for Excel export"
        
        # Save workbook
        workbook.save(excel_path)
        logging.info(f"Excel report saved to {excel_path}")
        
    except Exception as e:
        logging.error(f"Error creating Excel report: {str(e)}")

def create_summary_report(
    classification_results: List[Dict[str, Any]], 
    vendor_results: List[Dict[str, Any]], 
    employee_results: List[Dict[str, Any]], 
    failed_files: List[Dict[str, str]], 
    output_folder: str, 
    input_folder_name: str,
    total_files: int,
    elapsed_time: float,
    relevant_documents: List[Tuple[str, str]]
):
    """Create markdown summary report of the processing results."""
    summary_filename = f"{input_folder_name}_summary.md"
    summary_path = os.path.join(output_folder, summary_filename)
    
    try:
        # Calculate statistics
        classified_files = len(classification_results)
        relevant_files = len(relevant_documents)
        extracted_files = len(vendor_results) + len(employee_results)
        vendor_extracted = len(vendor_results)
        employee_extracted = len(employee_results)
        
        # Classification distribution
        classification_counts = {}
        page_stats = {"≤10": 0, "11-20": 0, ">20": 0}
        
        for result in classification_results:
            classification = result.get("classification", "unknown")
            classification_counts[classification] = classification_counts.get(classification, 0) + 1
            
            total_pages = result.get("total_pages_in_pdf", 0)
            if total_pages <= 10:
                page_stats["≤10"] += 1
            elif total_pages <= 20:
                page_stats["11-20"] += 1
            else:
                page_stats[">20"] += 1
        
        # Count oversized files that were processed
        oversized_processed = sum(1 for result in classification_results 
                                 if result.get("classification") in ["employee_t&e", "vendor_invoice"] 
                                 and result.get("total_pages_in_pdf", 0) > MAX_EXTRACTION_PAGES)
        
        # Count failed files by stage
        classification_failures = sum(1 for f in failed_files if f["failure_stage"] == "classification")
        extraction_failures = sum(1 for f in failed_files if f["failure_stage"] == "extraction")
        
        # Generate markdown content
        markdown_content = f"""# Invoice Processing Summary Report

## Processing Overview
- **Input Folder**: `{input_folder_name}`
- **Processing Date**: {time.strftime('%Y-%m-%d %H:%M:%S')}
- **Total Processing Time**: {elapsed_time:.2f} seconds
- **Configuration**: Enhanced 2-Step Pipeline with {CLASSIFICATION_MODEL} + {EXTRACTION_MODEL}

## File Processing Statistics

### Overall Results
| Metric | Count | Percentage |
|--------|-------|------------|
| **Total Files Processed** | {total_files} | 100.0% |
| **Successfully Classified** | {classified_files} | {classified_files/total_files*100:.1f}% |
| **Eligible for Extraction** | {relevant_files} | {relevant_files/total_files*100:.1f}% |
| **Successfully Extracted** | {extracted_files} | {extracted_files/relevant_files*100:.1f}% of eligible |
| **Failed Files** | {len(failed_files)} | {len(failed_files)/total_files*100:.1f}% |

### Classification Distribution
"""
        
        for classification, count in classification_counts.items():
            percentage = (count / classified_files) * 100 if classified_files > 0 else 0
            markdown_content += f"- **{classification.replace('_', ' ').title()}**: {count} files ({percentage:.1f}%)\n"
        
        markdown_content += f"""
### Page Distribution
"""
        for page_range, count in page_stats.items():
            percentage = (count / classified_files) * 100 if classified_files > 0 else 0
            markdown_content += f"- **{page_range} pages**: {count} files ({percentage:.1f}%)\n"
        
        markdown_content += f"""
### Extraction Results
- **Vendor Invoices Extracted**: {vendor_extracted}
- **Employee T&E Extracted**: {employee_extracted}
- **Oversized Files Processed** (using first {MAX_EXTRACTION_PAGES} pages): {oversized_processed}

## Processing Details

### Pipeline Configuration
- **Classification Model**: {CLASSIFICATION_MODEL}
- **Extraction Model**: {EXTRACTION_MODEL}
- **Classification Pages**: First {MAX_CLASSIFICATION_PAGES} pages analyzed
- **Extraction Page Limit**: ≤{MAX_EXTRACTION_PAGES} pages (truncated if larger)
- **Max Concurrent Classify**: {MAX_CONCURRENT_CLASSIFY}
- **Max Concurrent Extract**: {MAX_CONCURRENT_EXTRACT}

### Quality Metrics
- **Classification Success Rate**: {classified_files/total_files*100:.1f}%
- **Extraction Success Rate**: {extracted_files/relevant_files*100:.1f}% (of eligible files)
- **Overall Processing Success Rate**: {(total_files - len(failed_files))/total_files*100:.1f}%

"""
        
        if failed_files:
            markdown_content += f"""## Failed Files Analysis

### Failure Distribution
- **Classification Failures**: {classification_failures}
- **Extraction Failures**: {extraction_failures}

### Failed Files Details
"""
            for failure in failed_files:
                markdown_content += f"- **{failure['file_name']}**: {failure['failure_stage']} - {failure['error_message'][:100]}{'...' if len(failure['error_message']) > 100 else ''}\n"
        else:
            markdown_content += "## ✅ All Files Processed Successfully\nNo files failed during processing!\n"
        
        markdown_content += f"""
## Output Files Generated

### CSV Files
- `{CLASSIFICATION_CSV_FILE}` - Classification results for all files
- `{VENDOR_EXTRACTION_CSV_FILE}` - Vendor invoice extraction results
- `{EMPLOYEE_EXTRACTION_CSV_FILE}` - Employee T&E extraction results
"""
        
        if failed_files:
            markdown_content += f"- `{FAILED_CSV_FILE}` - Failed files list with error details\n"
        
        markdown_content += f"""
### Excel File
- `{input_folder_name}_data.xlsx` - Combined results with 3 worksheets:
  - Classification
  - Employee_T&E  
  - Vendor_Invoices

### Debug Information
- JSON responses saved to: `json_responses_2step_enhanced/{input_folder_name}/`
- Logs saved to: `{LOGS_FOLDER}/{LOG_FILE}`

---
*Report generated by Enhanced 2-Step Invoice Processing Pipeline*
"""
        
        # Write markdown file
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        
        logging.info(f"Summary report saved to {summary_path}")
        
    except Exception as e:
        logging.error(f"Error creating summary report: {str(e)}")

def create_chunked_summary_report(
    output_folder: str,
    input_folder_name: str,
    total_files: int,
    classified_files: int,
    relevant_files: int,
    extracted_files: int,
    vendor_extracted: int,
    employee_extracted: int,
    elapsed_time: float
):
    """Create simplified markdown summary report for chunked processing mode."""
    summary_filename = f"{input_folder_name}_summary.md"
    summary_path = os.path.join(output_folder, summary_filename)
    
    try:
        # Generate simplified markdown content for chunked mode
        # Safe percentage calculations
        classified_pct = (classified_files/total_files*100) if total_files > 0 else 0
        relevant_pct = (relevant_files/total_files*100) if total_files > 0 else 0
        extracted_pct = (extracted_files/relevant_files*100) if relevant_files > 0 else 0
        
        markdown_content = f"""# Invoice Processing Summary Report

## Processing Overview
- **Input Folder**: `{input_folder_name}`
- **Processing Date**: {time.strftime('%Y-%m-%d %H:%M:%S')}
- **Total Processing Time**: {elapsed_time:.2f} seconds
- **Configuration**: Enhanced 2-Step Pipeline with {CLASSIFICATION_MODEL} + {EXTRACTION_MODEL}
- **Processing Mode**: Chunked Processing (memory-efficient streaming)

## File Processing Statistics

### Overall Results
| Metric | Count | Percentage |
|--------|-------|------------|
| **Total Files Processed** | {total_files} | 100.0% |
| **Successfully Classified** | {classified_files} | {classified_pct:.1f}% |
| **Eligible for Extraction** | {relevant_files} | {relevant_pct:.1f}% |
| **Successfully Extracted** | {extracted_files} | {extracted_pct:.1f}% of eligible |

### Extraction Results
- **Vendor Invoices Extracted**: {vendor_extracted}
- **Employee T&E Extracted**: {employee_extracted}

## Processing Details

### Pipeline Configuration
- **Classification Model**: {CLASSIFICATION_MODEL}
- **Extraction Model**: {EXTRACTION_MODEL}
- **Classification Pages**: First {MAX_CLASSIFICATION_PAGES} pages analyzed
- **Extraction Page Limit**: ≤{MAX_EXTRACTION_PAGES} pages (truncated if larger)
- **Max Concurrent Classify**: {MAX_CONCURRENT_CLASSIFY}
- **Max Concurrent Extract**: {MAX_CONCURRENT_EXTRACT}
- **Processing Mode**: Chunked streaming (chunk size: {PROCESSING_CHUNK_SIZE})

### Quality Metrics
- **Classification Success Rate**: {classified_pct:.1f}%
- **Extraction Success Rate**: {extracted_pct:.1f}% (of eligible files)
- **Overall Processing Success Rate**: {(extracted_files/total_files*100) if total_files > 0 else 0:.1f}%

## Output Files Generated

### CSV Files
- `{CLASSIFICATION_CSV_FILE}` - Classification results for all files
- `{VENDOR_EXTRACTION_CSV_FILE}` - Vendor invoice extraction results
- `{EMPLOYEE_EXTRACTION_CSV_FILE}` - Employee T&E extraction results

### Processing Notes
- **Memory Optimization**: Used chunked processing with streaming CSV writers for large datasets
- **Detailed Statistics**: Use non-chunked mode (smaller datasets) for detailed breakdowns
- **Excel Reports**: Available in non-chunked mode; CSV files contain all data

### Debug Information
- JSON responses saved to: `json_responses_2step_enhanced/{input_folder_name}/`
- Logs saved to: `{LOGS_FOLDER}/{LOG_FILE}`

---
*Report generated by Enhanced 2-Step Invoice Processing Pipeline (Chunked Mode)*
"""
        
        # Write markdown file
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        
        logging.info(f"Chunked summary report saved to {summary_path}")
        
    except Exception as e:
        logging.error(f"Error creating chunked summary report: {str(e)}")

def create_dynamic_output_folders(input_folder: str, base_output_folder: str, base_json_folder: str) -> Tuple[str, str]:
    """Create dynamic output folders based on input folder name."""
    input_folder_name = os.path.basename(os.path.normpath(input_folder))
    
    # Create output folder with input folder name
    dynamic_output_folder = os.path.join(base_output_folder, input_folder_name)
    os.makedirs(dynamic_output_folder, exist_ok=True)
    
    # Create JSON responses folder with input folder name and subfolders for each extraction type
    dynamic_json_folder = os.path.join(base_json_folder, input_folder_name)
    os.makedirs(dynamic_json_folder, exist_ok=True)
    os.makedirs(os.path.join(dynamic_json_folder, "classification"), exist_ok=True)
    os.makedirs(os.path.join(dynamic_json_folder, "vendor_invoice"), exist_ok=True)
    os.makedirs(os.path.join(dynamic_json_folder, "employee_t&e"), exist_ok=True)
    
    return dynamic_output_folder, dynamic_json_folder

async def setup_processing_environment(
    input_folder: str, 
    base_output_folder: str, 
    logs_folder: str, 
    base_json_responses_folder: str,
    resume_extraction: bool = False
) -> Tuple[str, str, List[str], "genai.Client", CapacityLimiter]:
    """
    Set up the processing environment including folders, logging, file discovery, and client.
    
    Returns:
        Tuple of (output_folder, json_responses_folder, pdf_files, client, quota_semaphore)
    """
    # Create dynamic folder structure
    output_folder, json_responses_folder = create_dynamic_output_folders(
        input_folder, base_output_folder, base_json_responses_folder
    )
    
    # Ensure base directories exist
    os.makedirs(base_output_folder, exist_ok=True)
    os.makedirs(logs_folder, exist_ok=True)
    os.makedirs(base_json_responses_folder, exist_ok=True)
    
    # Set up logging using new config system
    setup_logging(Path(logs_folder), LOG_FILE)
    
    # Clear old response artifacts for this input folder (SKIP IN RESUME MODE)
    if not resume_extraction and os.path.exists(json_responses_folder):
        try:
            shutil.rmtree(json_responses_folder)
            logging.info(f"Cleared {json_responses_folder} directory.")
        except Exception as e:
            logging.error(f"Unable to clear {json_responses_folder}: {str(e)}")
    
    # Recreate folders with all subfolders
    os.makedirs(json_responses_folder, exist_ok=True)
    os.makedirs(os.path.join(json_responses_folder, "classification"), exist_ok=True)
    os.makedirs(os.path.join(json_responses_folder, "vendor_invoice"), exist_ok=True)
    os.makedirs(os.path.join(json_responses_folder, "employee_t&e"), exist_ok=True)
    
    # Handle resume extraction mode
    if resume_extraction:
        logging.info("🔄 Resume extraction mode enabled - finding missing extractions...")
        missing_files, total_eligible, total_completed = get_missing_extraction_files(output_folder, json_responses_folder)
        
        if not missing_files:
            logging.info("✅ All files have already been extracted! Nothing to resume.")
            return output_folder, json_responses_folder, [], None, None, None
        
        # Convert to format expected by extraction pipeline: List of (file_path, doc_type) tuples
        pdf_files = missing_files  # This will be handled differently in main()
        logging.info(f"📋 Resume mode: {len(missing_files)} files need extraction (completed: {total_completed}/{total_eligible})")
    else:
        # Normal mode: Find PDF files (including subdirectories)
        pdf_files = []
        for root, dirs, files in os.walk(input_folder):
            # Skip excluded directories
            dirs[:] = [d for d in dirs if d not in ["dup", "__pycache__"]]
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(root, file))
        
        if not pdf_files:
            logging.warning(f"No PDF files found in '{input_folder}'. Exiting.")
            return output_folder, json_responses_folder, [], None, None, None
    
    logging.info(f"🔍 Found {len(pdf_files)} PDFs to process in enhanced 2-step pipeline")
    logging.info(f"⚙️  Configuration: Classify={MAX_CONCURRENT_CLASSIFY}, Extract={MAX_CONCURRENT_EXTRACT} concurrent calls")
    logging.info(f"📄 Classification: {CLASSIFICATION_MODEL} (first {MAX_CLASSIFICATION_PAGES} pages)")
    logging.info(f"🤖 Extraction: {EXTRACTION_MODEL} (≤{MAX_EXTRACTION_PAGES} pages, first {MAX_EXTRACTION_PAGES} pages for oversized)")
    logging.info(f"📁 Output folder: {output_folder}")
    logging.info(f"🗂️  JSON responses: {json_responses_folder}")
    
    # Initialize genai client with aiohttp transport (Phase 3 optimization)
    http_options = types.HttpOptions(
        async_client_args={
            "connector": aiohttp.TCPConnector(limit=50, limit_per_host=10),
        }
    )
    
    if settings.use_vertex_ai:
        client = genai.Client(http_options=http_options)
        logging.info("🔐 Using Vertex AI with Application Default Credentials + aiohttp transport")
    else:
        client = genai.Client(api_key=API_KEY, http_options=http_options)
        logging.info("🔐 Using regular Gemini API with API key + aiohttp transport")
    
    # Create single quota semaphore for API rate limiting (Phase 1 optimization)
    quota_limiter = CapacityLimiter(QUOTA_LIMIT)
    
    return output_folder, json_responses_folder, pdf_files, client, quota_limiter

def get_missing_extraction_files(output_folder: str, json_responses_folder: str) -> Tuple[List[Tuple[str, str]], int, int]:
    """
    Compare classification results against existing extraction files to find missing ones.
    
    Args:
        output_folder: Path to output folder containing classification_results.csv
        json_responses_folder: Path to JSON responses folder containing extraction files
        
    Returns:
        Tuple of (missing_files_list, total_eligible, total_completed) where:
        - missing_files_list: List of (file_path, doc_type) tuples needing extraction
        - total_eligible: Total files eligible for extraction from classification
        - total_completed: Total files already extracted
    """
    import csv
    
    classification_csv = os.path.join(output_folder, CLASSIFICATION_CSV_FILE)
    if not os.path.exists(classification_csv):
        logging.error(f"Classification results not found: {classification_csv}")
        return [], 0, 0
    
    # Read classification results to get eligible files
    eligible_files = {}  # file_path -> doc_type
    with open(classification_csv, 'r', newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['Classification'] in ['vendor_invoice', 'employee_t&e']:
                eligible_files[row['File Path']] = row['Classification']
    
    total_eligible = len(eligible_files)
    logging.info(f"📋 Found {total_eligible} files eligible for extraction from classification results")
    
    # Check which files have already been extracted
    extracted_files = set()
    
    # Check vendor_invoice extractions (Phase 7: Fixed glob pattern)
    vendor_dir = os.path.join(json_responses_folder, "vendor_invoice")
    if os.path.exists(vendor_dir):
        import fnmatch
        for file_path in eligible_files:
            if eligible_files[file_path] == 'vendor_invoice':
                stem = Path(file_path).stem
                pattern = f"{stem}_extraction_vendor_invoice_attempt_*.txt"
                if any(fnmatch.filter(os.listdir(vendor_dir), pattern)):
                    extracted_files.add(file_path)
    
    # Check employee_t&e extractions (Phase 7: Fixed glob pattern)
    employee_dir = os.path.join(json_responses_folder, "employee_t&e")
    if os.path.exists(employee_dir):
        import fnmatch
        for file_path in eligible_files:
            if eligible_files[file_path] == 'employee_t&e':
                stem = Path(file_path).stem
                pattern = f"{stem}_extraction_employee_t&e_attempt_*.txt"
                if any(fnmatch.filter(os.listdir(employee_dir), pattern)):
                    extracted_files.add(file_path)
    
    total_completed = len(extracted_files)
    logging.info(f"✅ Found {total_completed} files already extracted")
    
    # Find missing files
    missing_files = []
    for file_path, doc_type in eligible_files.items():
        if file_path not in extracted_files:
            missing_files.append((file_path, doc_type))
    
    logging.info(f"🔄 Found {len(missing_files)} files needing extraction ({total_eligible - total_completed} missing)")
    
    return missing_files, total_eligible, total_completed

async def process_chunk_with_streaming_csv(
    pdf_chunk: List[str], 
    client: "genai.Client",
    quota_semaphore: asyncio.Semaphore,
    json_responses_folder: str,
    classification_writer: StreamingCSVWriter,
    vendor_writer: StreamingCSVWriter,
    employee_writer: StreamingCSVWriter,
    chunk_number: int,
    total_chunks: int
) -> Tuple[int, int, int, int]:
    """
    Process a chunk of PDF files with pipeline overlap and streaming CSV output.
    
    Returns:
        Tuple of (classified_count, extracted_count, failed_classify_count, failed_extract_count)
    """
    # Check for shutdown request before processing chunk
    global shutdown_event
    if shutdown_event and shutdown_event.is_set():
        logging.info(f"🛑 Shutdown requested, skipping chunk {chunk_number}")
        return 0, 0, 0, 0
    
    logging.info(f"📦 Processing chunk {chunk_number}/{total_chunks} ({len(pdf_chunk)} files)")
    
    # Process this chunk with pipeline overlap
    classification_results, extraction_results, failed_classification, failed_extraction, relevant_documents = await process_with_pipeline_overlap(
        pdf_chunk, client, quota_semaphore, json_responses_folder
    )
    
    # Stream classification results to CSV immediately
    classification_config = get_classification_csv_config()
    row_mapper = classification_config["row_mapper"]
    for result in classification_results:
        mapped_result = row_mapper(result)
        await classification_writer.write_row(mapped_result)
    
    # Stream extraction results to appropriate CSV files immediately (with flattening)
    for result in extraction_results:
        doc_type = result.get("document_type_processed", "")
        if doc_type == "vendor_invoice":
            for row in build_vendor_rows(result):
                await vendor_writer.write_row(row)
        elif doc_type == "employee_t&e":
            for row in build_employee_rows(result):
                await employee_writer.write_row(row)
    
    logging.info(f"✅ Chunk {chunk_number} complete: {len(classification_results)} classified, {len(extraction_results)} extracted")
    
    return len(classification_results), len(extraction_results), len(failed_classification), len(failed_extraction)

async def process_with_pipeline_overlap(
    pdf_files: List[str],
    client: "genai.Client",
    quota_semaphore: asyncio.Semaphore,
    json_responses_folder: str
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, str]], List[Dict[str, str]], List[Tuple[str, str]]]:
    """
    Process files with pipeline overlap: start extraction immediately as classification completes.
    
    Args:
        pdf_files: List of PDF file paths to process
        client: Shared genai client  
        quota_semaphore: Single semaphore controlling all API calls
        json_responses_folder: Folder to save JSON responses
        
    Returns:
        Tuple of (classification_results, extraction_results, failed_classification, failed_extraction, relevant_documents)
    """
    classification_results = []
    extraction_results = []
    failed_classification = []
    failed_extraction = []
    relevant_documents = []
    
    # Use asyncio.Queue to stream classification results to extraction
    extraction_queue = asyncio.Queue()
    classification_done = asyncio.Event()
    
    async def classification_worker():
        """Process all classifications and put relevant results in extraction queue"""
        nonlocal classification_results, failed_classification
        
        # Process classification with retries using existing infrastructure 
        classify_results, classify_failed = await process_with_retries(
            pdf_files, client, quota_semaphore, json_responses_folder,
            stage="classification", max_passes=3
        )
        
        classification_results.extend(classify_results)
        failed_classification.extend(classify_failed)
        
        # Queue relevant documents for extraction
        for result in classify_results:
            classification = result.get("classification", "")
            if classification in ["employee_t&e", "vendor_invoice"]:
                await extraction_queue.put((result["file_path"], classification))
        
        # Signal that classification is done
        await extraction_queue.put(None)  # Sentinel value
        classification_done.set()
        
    async def extraction_worker():
        """Process extractions as they become available from classification"""
        nonlocal extraction_results, failed_extraction, relevant_documents
        batch_documents = []
        
        while True:
            item = await extraction_queue.get()
            if item is None:  # Sentinel - classification is done
                break
            relevant_documents.append(item)
            batch_documents.append(item)
            
            # Process in batches to avoid overwhelming the extraction pipeline
            if len(batch_documents) >= 50:  # Process in batches of 50
                extract_results, extract_failed = await process_with_retries(
                    batch_documents, client, quota_semaphore, json_responses_folder,
                    stage="extraction", max_passes=3
                )
                extraction_results.extend(extract_results)
                failed_extraction.extend(extract_failed)
                batch_documents = []
        
        # Process any remaining documents
        if batch_documents:
            extract_results, extract_failed = await process_with_retries(
                batch_documents, client, quota_semaphore, json_responses_folder,
                stage="extraction", max_passes=3
            )
            extraction_results.extend(extract_results)
            failed_extraction.extend(extract_failed)
    
    # Start both workers concurrently for pipeline overlap
    await asyncio.gather(
        classification_worker(),
        extraction_worker()
    )
    
    return classification_results, extraction_results, failed_classification, failed_extraction, relevant_documents

async def main(input_folder=None, output_folder=None, logs_folder=None, json_responses_folder=None, resume_extraction=False, resume=False, no_tui=False):
    """
    Enhanced main 2-step processing function with improved accuracy and organization.
    Step 1: Classify documents using gemini-2.5-flash (first 7 pages)
    Step 2: Extract data from relevant documents ≤20 pages using gemini-2.5-pro with enhanced prompts
    """
    # Track execution time for summary report
    start_time = time.time()
    
    # Use provided arguments or fall back to defaults
    input_folder = input_folder or INPUT_FOLDER
    base_output_folder = output_folder or OUTPUT_FOLDER
    logs_folder = logs_folder or LOGS_FOLDER
    base_json_responses_folder = json_responses_folder or JSON_RESPONSES_FOLDER
    
    # Set up processing environment
    output_folder, json_responses_folder, pdf_files, client, quota_limiter = await setup_processing_environment(
        input_folder, base_output_folder, logs_folder, base_json_responses_folder, resume_extraction
    )
    
    # Initialize PDF file descriptor semaphore (Phase 4 optimization)
    global pdf_fd_semaphore
    pdf_fd_semaphore = CapacityLimiter(PDF_FD_SEMAPHORE_LIMIT)
    logging.info(f"🔒 PDF file descriptor semaphore initialized (limit: {PDF_FD_SEMAPHORE_LIMIT})")
    
    # Initialize processing manifest for resumable operations
    global processing_manifest, pause_event, shutdown_event
    if resume:
        manifest_path = os.path.join(base_output_folder, "manifest.db")
        processing_manifest = ProcessingManifest(manifest_path)
        processing_manifest.connect()
        logging.info(f"📋 SQLite manifest initialized for resumable processing: {manifest_path}")
        
        # Initialize TUI for interactive monitoring
        total_files = len(pdf_files) if pdf_files else 0
        tui_result = await run_tui_monitor(processing_manifest, total_files, disable_tui=no_tui)
        
        if len(tui_result) == 3:  # TUI enabled
            pause_event, shutdown_event, tui_task = tui_result
        else:  # TUI disabled
            pause_event, shutdown_event = tui_result
            tui_task = None
            
        logging.info(f"🎛️  TUI and pause/resume controls initialized (TUI enabled: {not no_tui})")
    else:
        processing_manifest = None
        pause_event = None
        shutdown_event = None
    
    # Check if setup was successful
    if not pdf_files or client is None:
        return
    
    if resume and processing_manifest:
        # ===========================
        # SQLITE MANIFEST RESUME MODE: USE MANIFEST TO DETERMINE RESUME QUEUES
        # ===========================
        logging.info("🔄 SQLite manifest resume mode: Determining files to process...")
        
        # Get all PDF files from the input folder
        all_pdf_files = []
        for root, dirs, files in os.walk(input_folder):
            for file in files:
                if file.lower().endswith('.pdf'):
                    all_pdf_files.append(os.path.join(root, file))
        
        # Use manifest to determine what needs to be processed
        classify_list, extract_list = processing_manifest.get_resume_queues(all_pdf_files)
        
        logging.info(f"📋 Resume queues: {len(classify_list)} need classification, {len(extract_list)} need extraction")
        
        # Set up for regular processing with filtered lists
        pdf_files = all_pdf_files  # Will be filtered by manifest during processing
        
        # Initialize result variables for manifest mode
        classification_results = []
        extraction_results = []
        failed_classification = []
        failed_extraction = []
        
    elif resume_extraction:
        # ===========================
        # RESUME MODE: SKIP CLASSIFICATION, USE EXISTING RESULTS
        # ===========================
        logging.info("🔄 Resume mode: Skipping classification, using existing results...")
        
        # In resume mode, pdf_files contains (file_path, doc_type) tuples
        relevant_documents = pdf_files  # Already in the right format
        classification_results = []  # Not needed for resume mode
        failed_classification = []   # Not needed for resume mode
        extraction_results = []      # Will be populated by extraction process
        failed_extraction = []       # Will be populated by extraction process
        
        logging.info(f"📋 Loaded {len(relevant_documents)} files for extraction from previous classification")
        
        # Skip to extraction logic
        # TODO: Handle extraction with manifest
        
    if not resume or not processing_manifest:
        # ===========================
        # PHASE 6: CHUNKED PROCESSING WITH STREAMING CSV WRITERS
        # ===========================
        
        total_files = len(pdf_files)
        chunk_size = PROCESSING_CHUNK_SIZE
        chunks = list(chunker(pdf_files, chunk_size))
        total_chunks = len(chunks)
        
        logging.info(f"📦 Starting Chunked Pipeline Processing ({total_files} files, {total_chunks} chunks of {chunk_size})")
        logging.info("🏷️  Using streaming CSV writers for memory efficiency...")
        
        # Initialize streaming CSV writers
        classification_csv_path = os.path.join(output_folder, CLASSIFICATION_CSV_FILE)
        vendor_csv_path = os.path.join(output_folder, VENDOR_EXTRACTION_CSV_FILE)
        employee_csv_path = os.path.join(output_folder, EMPLOYEE_EXTRACTION_CSV_FILE)
        
        classification_writer = StreamingCSVWriter(classification_csv_path, get_classification_csv_config()["fieldnames"])
        vendor_writer = StreamingCSVWriter(vendor_csv_path, get_vendor_extraction_csv_config()["fieldnames"])
        employee_writer = StreamingCSVWriter(employee_csv_path, get_employee_extraction_csv_config()["fieldnames"])
        
        # Process chunks with garbage collection between chunks
        total_classified = total_extracted = total_failed_classify = total_failed_extract = 0
        
        try:
            for chunk_num, chunk in enumerate(chunks, 1):
                # Check for shutdown request before processing each chunk
                if shutdown_event and shutdown_event.is_set():
                    logging.info(f"🛑 Shutdown requested, stopping after chunk {chunk_num-1}/{total_chunks}")
                    break
                
                classified, extracted, failed_classify, failed_extract = await process_chunk_with_streaming_csv(
                    chunk, client, quota_semaphore, json_responses_folder,
                    classification_writer, vendor_writer, employee_writer,
                    chunk_num, total_chunks
                )
                
                total_classified += classified
                total_extracted += extracted
                total_failed_classify += failed_classify
                total_failed_extract += failed_extract
                
                # Force garbage collection between chunks to free memory
                if chunk_num < total_chunks:
                    gc.collect()
                    logging.info(f"🗑️  Memory cleanup after chunk {chunk_num}/{total_chunks}")
        
        finally:
            # Close all streaming writers
            classification_writer.close()
            vendor_writer.close() 
            employee_writer.close()
            
        logging.info(f"📊 Chunked processing complete: {total_classified} classified, {total_extracted} extracted")
        # Track vendor / employee extraction counts from streaming writers
        chunked_vendor_extracted = vendor_writer.rows_written
        chunked_employee_extracted = employee_writer.rows_written
        
        # Initialize empty result arrays for compatibility with final reporting
        classification_results = []
        extraction_results = []
        failed_classification = []
        failed_extraction = []
    
    if not resume_extraction:
        # Enhanced classification summary (only for non-resume mode and non-chunked mode)
        if classification_results:  # Only if we have detailed classification data
            classification_counts = {}
            page_stats = {"≤10": 0, "11-20": 0, ">20": 0}
            
            for result in classification_results:
                classification = result.get("classification", "unknown")
                classification_counts[classification] = classification_counts.get(classification, 0) + 1
                
                total_pages = result.get("total_pages_in_pdf", 0)
                if total_pages <= 10:
                    page_stats["≤10"] += 1
                elif total_pages <= 20:
                    page_stats["11-20"] += 1
                else:
                    page_stats[">20"] += 1
            
            logging.info(f"📊 Enhanced Classification Results:")
            total_classified = len(classification_results)
            for classification, count in classification_counts.items():
                percentage = (count / total_classified) * 100 if total_classified > 0 else 0
                logging.info(f"   {classification}: {count} files ({percentage:.1f}%)")
            
            logging.info(f"📖 Page Distribution:")
            for page_range, count in page_stats.items():
                percentage = (count / total_classified) * 100 if total_classified > 0 else 0
                logging.info(f"   {page_range} pages: {count} files ({percentage:.1f}%)")
        else:
            logging.info(f"📊 Enhanced Classification Results: Detailed breakdown skipped in chunked processing mode")
    
    # ===========================
    # EXTRACTION RESULTS HANDLING
    # ===========================
    
    if resume_extraction:
        # Resume mode - process missing files
        logging.info(f"💎 Step 2: Enhanced Extraction from {len(relevant_documents)} eligible documents")
        logging.info(f"🔄 Resume mode: Processing {len(relevant_documents)} files needing extraction")
        
        if not relevant_documents:
            logging.info("No eligible documents found for extraction. Process complete.")
            return
        
        extraction_results, failed_extraction = await process_with_retries(
            relevant_documents, client, quota_semaphore, json_responses_folder,
            stage="extraction", max_passes=3
        )
    else:
        # Pipeline overlap mode - extraction_results already available
        # Show processing statistics based on classification results
        if classification_results:
            # Non-chunked mode: calculate from in-memory classification results
            skipped_irrelevant = len([r for r in classification_results if r.get("classification") == "irrelevant"])
            skipped_processing_failed = len([r for r in classification_results if r.get("classification") == "processing_failed"]) 
            oversized_count = len([r for r in classification_results if r.get("total_pages_in_pdf", 0) > MAX_EXTRACTION_PAGES and r.get("classification") in ["employee_t&e", "vendor_invoice"]])
            relevant_count = len([r for r in classification_results if r.get("classification") in ["employee_t&e", "vendor_invoice"]])
            
            # Create relevant_documents list from classification results for reporting
            relevant_documents = [
                (r["file_path"], r.get("classification"))
                for r in classification_results 
                if r.get("classification") in ["employee_t&e", "vendor_invoice"]
            ]
        else:
            # Chunked mode: calculate from streaming CSV file since classification_results is empty to save memory
            import csv
            classification_csv_path = os.path.join(output_folder, CLASSIFICATION_CSV_FILE)
            skipped_irrelevant = 0
            skipped_processing_failed = 0
            oversized_count = 0
            relevant_count = 0
            relevant_documents = []
            
            if os.path.exists(classification_csv_path):
                with open(classification_csv_path, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        classification = row.get('Classification', '')
                        total_pages = int(row.get('Total Pages In PDF', 0)) if row.get('Total Pages In PDF', '').isdigit() else 0
                        file_path = row.get('File Path', '')
                        
                        if classification == "irrelevant":
                            skipped_irrelevant += 1
                        elif classification == "processing_failed":
                            skipped_processing_failed += 1
                        elif classification in ["employee_t&e", "vendor_invoice"]:
                            relevant_count += 1
                            relevant_documents.append((file_path, classification))
                            if total_pages > MAX_EXTRACTION_PAGES:
                                oversized_count += 1
        
        # Adjust logging based on processing mode (chunked vs. non-chunked)
        if classification_results:
            logging.info(f"💎 Pipeline Processing Complete - {len(extraction_results)} documents extracted")
        else:
            extracted_for_log = total_extracted if 'total_extracted' in locals() else len(extraction_results)
            logging.info(f"💎 Pipeline Processing Complete - {extracted_for_log} documents extracted")
        logging.info(f"⚡ Processing Statistics:")
        logging.info(f"   📋 Irrelevant documents skipped: {skipped_irrelevant}")
        logging.info(f"   💥 Processing failed documents skipped: {skipped_processing_failed}")
        logging.info(f"   📚 Oversized files (using first {MAX_EXTRACTION_PAGES} pages): {oversized_count}")
        logging.info(f"   💰 Total relevant documents processed: {relevant_count} files")
        
        # Log oversized files for reference
        for result in classification_results:
            classification = result.get("classification", "")
            total_pages = result.get("total_pages_in_pdf", 0)
            file_path = result["file_path"]
            if classification in ["employee_t&e", "vendor_invoice"] and total_pages > MAX_EXTRACTION_PAGES:
                logging.info(f"[PROCESS] {os.path.basename(file_path)} - Used first {MAX_EXTRACTION_PAGES} pages ({total_pages} total pages)")
    
    # Note: Retries are now handled automatically within process_with_retries()
    
    # ===========================
    # SAVE FINAL RESULTS TO CSV
    # ===========================
    
    # Separate results by document type for separate CSV files
    vendor_results = [r for r in extraction_results if r.get("document_type_processed") == "vendor_invoice"]
    employee_results = [r for r in extraction_results if r.get("document_type_processed") == "employee_t&e"]
    
    # Save separate extraction results
    vendor_csv_path = os.path.join(output_folder, VENDOR_EXTRACTION_CSV_FILE)
    employee_csv_path = os.path.join(output_folder, EMPLOYEE_EXTRACTION_CSV_FILE)
    
    save_vendor_extraction_results_to_csv(vendor_results, vendor_csv_path)
    save_employee_extraction_results_to_csv(employee_results, employee_csv_path)
    
    # ===========================
    # ENHANCED FINAL SUMMARY
    # ===========================
    
    # Handle both chunked and non-chunked modes for statistics
    if not classification_results and 'total_classified' in locals():
        # Chunked mode - use the accumulated counts
        total_files = len(pdf_files)
        classified_files = total_classified
        extracted_files = total_extracted
        vendor_extracted = chunked_vendor_extracted if 'chunked_vendor_extracted' in locals() else 0
        employee_extracted = chunked_employee_extracted if 'chunked_employee_extracted' in locals() else 0
        relevant_files = len(relevant_documents) if relevant_documents else extracted_files
    else:
        # Non-chunked mode - use the arrays
        total_files = len(pdf_files)
        classified_files = len(classification_results)
        relevant_files = len(relevant_documents)
        extracted_files = len(extraction_results)
        vendor_extracted = len(vendor_results)
        employee_extracted = len(employee_results)
    
    # Count oversized files that were processed
    oversized_processed = sum(1 for result in classification_results 
                             if result.get("classification") in ["employee_t&e", "vendor_invoice"] 
                             and result.get("total_pages_in_pdf", 0) > MAX_EXTRACTION_PAGES)
    
    logging.info(f"🎯 ENHANCED 2-STEP FINAL RESULTS:")
    logging.info(f"   📁 Total files processed: {total_files}")
    
    # Safe percentage calculations to avoid division by zero
    classified_pct = (classified_files/total_files*100) if total_files > 0 else 0
    relevant_pct = (relevant_files/total_files*100) if total_files > 0 else 0
    extracted_pct = (extracted_files/relevant_files*100) if relevant_files > 0 else 0
    
    logging.info(f"   🏷️  Successfully classified: {classified_files} ({classified_pct:.1f}%)")
    logging.info(f"   💎 Eligible for extraction: {relevant_files} ({relevant_pct:.1f}%)")
    logging.info(f"   ✅ Successfully extracted: {extracted_files} ({extracted_pct:.1f}% of eligible)")
    logging.info(f"   🏢 Vendor invoices extracted: {vendor_extracted}")
    logging.info(f"   👤 Employee T&E extracted: {employee_extracted}")
    logging.info(f"   ⚡ Total non-processable files skipped: {total_files - relevant_files} files (irrelevant + processing failed)")
    logging.info(f"   📚 Oversized files processed with first {MAX_EXTRACTION_PAGES} pages: {oversized_processed}")
    
    # Save failed files (combine both classification and extraction failures)
    # Filter out files that were actually successfully processed
    successfully_processed_files = set()
    for result in classification_results:
        successfully_processed_files.add(result["file_path"])
    for result in extraction_results:
        successfully_processed_files.add(result["file_path"])
    
    all_failed_files = failed_classification + failed_extraction
    # Remove duplicates based on file_path and exclude successfully processed files
    seen_files = set()
    unique_failed_files = []
    for failure in all_failed_files:
        file_path = failure["file_path"]
        if file_path not in seen_files and file_path not in successfully_processed_files:
            unique_failed_files.append(failure)
            seen_files.add(file_path)
    
    if unique_failed_files:
        failed_csv_path = os.path.join(output_folder, FAILED_CSV_FILE)
        save_results_to_csv(
            unique_failed_files, 
            failed_csv_path, 
            get_failed_files_csv_config(),
            "failed files"
        )
        logging.warning(f"📋 {len(unique_failed_files)} files failed processing. See {failed_csv_path}")
    else:
        logging.info("🎉 All eligible files processed successfully!")
    
    # ===========================
    # GENERATE EXCEL AND SUMMARY REPORTS
    # ===========================
    
    # Get input folder name for file naming
    input_folder_name = os.path.basename(os.path.normpath(input_folder))
    
    # Create Excel and summary reports (only if we have data arrays or adjust for chunked mode)
    if classification_results or vendor_results or employee_results:
        # Create Excel report with 3 worksheets
        logging.info("📊 Creating Excel report with 3 worksheets...")
        create_excel_report(
            classification_results, 
            vendor_results, 
            employee_results, 
            output_folder, 
            input_folder_name
        )
        
        # Create markdown summary report
        logging.info("📝 Creating summary report...")
        # Calculate elapsed time
        elapsed_time = time.time() - start_time
        create_summary_report(
            classification_results,
            vendor_results, 
            employee_results,
            unique_failed_files,
            output_folder,
            input_folder_name,
            total_files,
            elapsed_time,
            relevant_documents
        )
    else:
        # Chunked mode - create simplified reports since arrays are empty for memory efficiency
        logging.info("📊 Chunked mode: Creating simplified summary report...")
        elapsed_time = time.time() - start_time
        # Create simplified markdown summary
        create_chunked_summary_report(
            output_folder,
            input_folder_name,
            total_files,
            classified_files,
            relevant_files,
            extracted_files,
            vendor_extracted,
            employee_extracted,
            elapsed_time
        )
        # Also generate an Excel workbook from streamed CSVs
        logging.info("📊 Creating Excel workbook from streamed CSVs (chunked mode)...")
        create_excel_from_streamed_csvs(output_folder, input_folder_name)
    
    logging.info(f"📊 Enhanced results saved to {output_folder}/")
    logging.info(f"   - Classification: {CLASSIFICATION_CSV_FILE}")
    logging.info(f"   - Vendor Extraction: {VENDOR_EXTRACTION_CSV_FILE}")
    logging.info(f"   - Employee Extraction: {EMPLOYEE_EXTRACTION_CSV_FILE}")
    logging.info(f"   - Excel Report: {input_folder_name}_data.xlsx")
    logging.info(f"   - Summary Report: {input_folder_name}_summary.md")
    
    # ===========================
    # CLEANUP TUI AND MANIFEST
    # ===========================
    
    # Clean up TUI task if it was started
    if 'tui_task' in locals() and tui_task:
        try:
            tui_task.cancel()
            await tui_task
        except asyncio.CancelledError:
            pass  # Expected when cancelling
        except Exception as e:
            logging.debug(f"Error cleaning up TUI task: {e}")
    
    # Close manifest connection
    if processing_manifest:
        try:
            processing_manifest.close()
            logging.debug("📋 Manifest connection closed")
        except Exception as e:
            logging.debug(f"Error closing manifest: {e}")
    
    # Memory cleanup - clear result lists after all reporting is complete
    classification_results = []
    extraction_results = []
    failed_classification = []
    failed_extraction = []

if __name__ == "__main__":
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Enhanced 2-Step PDF processing: Classification then Specialized Extraction')
    parser.add_argument('--input', default=INPUT_FOLDER,
                       help=f'Input folder containing PDF files (default: {INPUT_FOLDER})')
    parser.add_argument('--batch', nargs='+', 
                       help='Process multiple input folders in sequence (e.g., --batch folder1 folder2 folder3)')
    parser.add_argument('--resume-extraction', action='store_true',
                       help='Resume extraction for files that failed/missed in previous run (skips classification)')
    parser.add_argument('--resume', action='store_true',
                       help='Resume processing using SQLite manifest (both classification and extraction)')
    parser.add_argument('--no-tui', action='store_true',
                       help='Disable interactive TUI for headless mode')
    parser.add_argument('--output', default=OUTPUT_FOLDER,
                       help=f'Output folder for results (default: {OUTPUT_FOLDER})')
    parser.add_argument('--logs', default=LOGS_FOLDER,
                       help=f'Logs folder (default: {LOGS_FOLDER})')
    parser.add_argument('--json-responses', default=JSON_RESPONSES_FOLDER,
                       help=f'JSON responses folder (default: {JSON_RESPONSES_FOLDER})')
    
    args = parser.parse_args()
    
    # Set up basic logging
    setup_logging(Path(args.logs), LOG_FILE)
    
    # Log configuration
    use_vertex = settings.use_vertex_ai
    if use_vertex:
        project = settings.google_cloud_project
        location = settings.google_cloud_location
        logging.info(f"🚀 Starting Enhanced 2-Step Pipeline with Vertex AI - Project: {project}, Location: {location}")
    else:
        logging.info("🚀 Starting Enhanced 2-Step Pipeline with regular Gemini API")
    
    # Handle batch processing
    if args.batch:
        input_folders = args.batch
        logging.info(f"🔄 Batch mode: Processing {len(input_folders)} folders")
    else:
        input_folders = [args.input]
        logging.info(f"📂 Single folder mode: {args.input}")
    
    logging.info(f"📁 Base output folder: {args.output}")
    logging.info(f"📋 Logs folder: {args.logs}")
    logging.info(f"🗂️  Base JSON responses folder: {args.json_responses}")
    logging.info(f"🔧 Enhanced features: 7-page classification, ≤{MAX_EXTRACTION_PAGES}-page extraction, separate CSVs, dynamic folders")
    
    total_start_time = time.time()
    
    # Process each folder
    for i, folder in enumerate(input_folders, 1):
        logging.info(f"📁 Processing folder {i}/{len(input_folders)}: {folder}")
        folder_start_time = time.time()
        
        try:
            asyncio.run(main(
                input_folder=folder,
                output_folder=args.output,
                logs_folder=args.logs,
                json_responses_folder=args.json_responses,
                resume_extraction=args.resume_extraction,
                resume=args.resume,
                no_tui=args.no_tui
            ))
            folder_elapsed = time.time() - folder_start_time
            logging.info(f"✅ Folder {i}/{len(input_folders)} completed in {folder_elapsed:.2f} seconds: {folder}")
        except Exception as e:
            folder_elapsed = time.time() - folder_start_time
            # Handle coroutine errors properly
            error_msg = str(e)
            if "coroutine" in error_msg and hasattr(e, '__cause__') and e.__cause__:
                error_msg = str(e.__cause__)
            elif len(error_msg) > 200:
                error_msg = error_msg[:200] + "..."
            logging.error(f"❌ Folder {i}/{len(input_folders)} failed after {folder_elapsed:.2f} seconds: {folder} - {error_msg}")
    
    total_elapsed = time.time() - total_start_time
    logging.info(f"⏱️  Total batch execution time: {total_elapsed:.2f} seconds for {len(input_folders)} folders")