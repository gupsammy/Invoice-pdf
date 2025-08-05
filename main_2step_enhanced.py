import os
import json
import csv
import time
import logging
import shutil
import asyncio
import argparse
import fitz  # PyMuPDF
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
from dotenv import load_dotenv
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill

# Import Google Generative AI
from google import genai
from google.genai import types

# Import enhanced 2-step prompts v2
from prompts_2step_enhanced_v2 import (
    ENHANCED_CLASSIFICATION_PROMPT_V2,
    ENHANCED_EMPLOYEE_REIMBURSEMENT_EXTRACTION_PROMPT_V2,
    ENHANCED_VENDOR_INVOICE_EXTRACTION_PROMPT_V2
)

# ===========================
# ENHANCED CONFIGURATION CONSTANTS
# ===========================

# Folder Configuration
INPUT_FOLDER = "input"
OUTPUT_FOLDER = "output"
LOGS_FOLDER = "logs"
JSON_RESPONSES_FOLDER = "json_responses"

# File Configuration
CLASSIFICATION_CSV_FILE = "classification_results.csv"
VENDOR_EXTRACTION_CSV_FILE = "vendor_extraction_results.csv"
EMPLOYEE_EXTRACTION_CSV_FILE = "employee_extraction_results.csv"
COMBINED_CSV_FILE = "combined_enhanced_results.csv"
FAILED_CSV_FILE = "failed_enhanced_files.csv"
LOG_FILE = "invoice_extraction_2step_enhanced.log"

# Processing Configuration
_MAX_RETRIES = 3
_RETRY_DELAY_BASE_SECONDS = 10
MAX_CONCURRENT_API_CALLS = 5

# Enhanced 2-Step Configuration
MAX_CLASSIFICATION_PAGES = 7  # Increased from 5 to 7 pages
MAX_EXTRACTION_PAGES = 20     # Only process files with ≤20 pages for extraction (split if >20)
CLASSIFICATION_MODEL = "gemini-2.5-flash"
EXTRACTION_MODEL = "gemini-2.5-pro"

def setup_logging(logs_folder):
    """Set up logging with the specified logs folder."""
    os.makedirs(logs_folder, exist_ok=True)
    
    # Clear any existing handlers
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(os.path.join(logs_folder, LOG_FILE)),
            logging.StreamHandler()
        ]
    )

# Load environment variables
load_dotenv()
API_KEY = os.getenv("GEMINI_API_KEY")
if not API_KEY:
    raise ValueError("GEMINI_API_KEY not found in environment variables. Please set it in a .env file.")

def get_pdf_page_count(pdf_path: str) -> int:
    """Get the total number of pages in a PDF file."""
    try:
        doc = fitz.open(pdf_path)
        page_count = len(doc)
        doc.close()
        return page_count
    except Exception as e:
        logging.error(f"Error getting page count for {pdf_path}: {str(e)}")
        return 0

def extract_first_n_pages_pdf(pdf_path: str, max_pages: int = MAX_CLASSIFICATION_PAGES) -> bytes:
    """
    Extract the first N pages from a PDF and return as bytes.
    
    Args:
        pdf_path: Path to the PDF file
        max_pages: Maximum number of pages to extract (default: 7)
        
    Returns:
        PDF bytes containing only the first N pages
    """
    try:
        # Open the source PDF
        source_doc = fitz.open(pdf_path)
        
        # Create a new PDF with only the first N pages
        new_doc = fitz.open()
        
        # Copy the first N pages (or all pages if fewer than N)
        pages_to_copy = min(len(source_doc), max_pages)
        for page_num in range(pages_to_copy):
            new_doc.insert_pdf(source_doc, from_page=page_num, to_page=page_num)
        
        # Convert to bytes
        pdf_bytes = new_doc.tobytes()
        
        # Close documents
        source_doc.close()
        new_doc.close()
        
        return pdf_bytes
        
    except Exception as e:
        logging.error(f"Error extracting first {max_pages} pages from {pdf_path}: {str(e)}")
        raise

def create_preprocessing_failure_result(pdf_path: str, error_message: str) -> Dict[str, Any]:
    """
    Create a classification result for preprocessing failures.
    Mark as processing_failed without sending to API.
    """
    pdf_path_obj = Path(pdf_path)
    return {
        "file_name": pdf_path_obj.name,
        "file_path": pdf_path,
        "classification": "processing_failed",
        "confidence": 1.0,
        "reasoning": f"Preprocessing failed: {error_message}",
        "key_indicators": "PDF preprocessing error",
        "has_employee_codes": False,
        "has_vendor_letterhead": False,
        "has_invoice_numbers": False,
        "has_travel_dates": False,
        "appears_financial": False,
        "has_amount_calculations": False,
        "has_tax_information": False,
        "contains_multiple_doc_types": False,
        "primary_document_type": "preprocessing_error",
        "classification_model": "preprocessing-error-handler",
        "total_pages_in_pdf": 0,
        "pages_analyzed": 0,
        "classification_notes": f"File failed PDF preprocessing: {error_message}",
        "preprocessing_failure": True,
        "failure_stage": "preprocessing",
        "error_message": error_message
    }

async def classify_document_async(
    pdf_path: str,
    client: "genai.Client",
    semaphore: asyncio.Semaphore,
    json_responses_folder: str
) -> Optional[Dict[str, Any]]:
    """
    Classify a document using the first 7 pages and enhanced classification prompt.
    
    Args:
        pdf_path: Path to the PDF file
        client: Shared genai client
        semaphore: Semaphore to control concurrent API calls
        json_responses_folder: Folder to save JSON responses
        
    Returns:
        Dictionary containing classification results or None if failed
    """
    pdf_path_obj = Path(pdf_path)
    
    try:
        # Get total page count - catch preprocessing errors
        try:
            total_pages = get_pdf_page_count(pdf_path)
            if total_pages == 0:
                error_msg = "Unable to read PDF or empty PDF"
                logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Preprocessing error: {error_msg}")
                return create_preprocessing_failure_result(pdf_path, error_msg)
        except Exception as e:
            error_msg = f"Failed to get PDF page count: {str(e)}"
            logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Preprocessing error: {error_msg}")
            return create_preprocessing_failure_result(pdf_path, error_msg)
        
        # Extract first 7 pages - catch preprocessing errors
        try:
            pdf_bytes = extract_first_n_pages_pdf(pdf_path, MAX_CLASSIFICATION_PAGES)
        except Exception as e:
            error_msg = f"Failed to extract PDF pages: {str(e)}"
            logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Preprocessing error: {error_msg}")
            return create_preprocessing_failure_result(pdf_path, error_msg)
        
        contents = [
            types.Part.from_bytes(
                data=pdf_bytes,
                mime_type='application/pdf',
            ),
            f"Classify this document into one of the three categories based on the first {MAX_CLASSIFICATION_PAGES} pages."
        ]
        
        config = types.GenerateContentConfig(system_instruction=ENHANCED_CLASSIFICATION_PROMPT_V2)
        
        async with semaphore:
            for attempt in range(_MAX_RETRIES):
                try:
                    logging.info(f"[CLASSIFY] {pdf_path_obj.name} - Attempt {attempt + 1}/{_MAX_RETRIES} (Total pages: {total_pages})")
                    
                    response = await client.aio.models.generate_content(
                        model=CLASSIFICATION_MODEL,
                        contents=contents,
                        config=config
                    )
                    
                    # Save response for debugging
                    classification_log_filename = f"{pdf_path_obj.stem}_classification_attempt_{attempt+1}.txt"
                    classification_log_path = os.path.join(json_responses_folder, "classification", classification_log_filename)
                    os.makedirs(os.path.dirname(classification_log_path), exist_ok=True)
                    with open(classification_log_path, 'w', encoding='utf-8') as f:
                        f.write(response.text)
                    
                    # Parse JSON response with robust error handling
                    resp_txt = response.text
                    json_start = resp_txt.find('{')
                    json_end = resp_txt.rfind('}') + 1
                    
                    if json_start == -1 or json_end == 0:
                        if any(err in resp_txt for err in ["502", "Service Unavailable", "server error"]):
                            raise RuntimeError("Transient upstream error – will retry")
                        logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Unparsable response: {resp_txt[:120]}")
                        return None
                    
                    try:
                        classification_data = json.loads(resp_txt[json_start:json_end])
                    except json.JSONDecodeError as json_error:
                        # Try to fix common JSON malformations
                        logging.warning(f"[CLASSIFY] {pdf_path_obj.name} - JSON malformed, attempting repair: {str(json_error)}")
                        try:
                            import re
                            json_str = resp_txt[json_start:json_end]
                            # Fix common issue: "text" (extra text) -> "text"
                            fixed_json = re.sub(r'"([^"]*)" \([^)]*\)', r'"\1"', json_str)
                            classification_data = json.loads(fixed_json)
                            logging.info(f"[CLASSIFY] {pdf_path_obj.name} - JSON repair successful")
                        except json.JSONDecodeError:
                            logging.error(f"[CLASSIFY] {pdf_path_obj.name} - JSON repair failed")
                            raise json_error
                    
                    # Add metadata
                    classification_data["file_name"] = pdf_path_obj.name
                    classification_data["file_path"] = str(pdf_path_obj)
                    classification_data["classification_model"] = CLASSIFICATION_MODEL
                    classification_data["total_pages_in_pdf"] = total_pages
                    classification_data["pages_analyzed"] = min(total_pages, MAX_CLASSIFICATION_PAGES)
                    
                    classification = classification_data.get('classification', 'unknown')
                    confidence = classification_data.get('confidence', 0)
                    logging.info(f"[CLASSIFY] {pdf_path_obj.name} - Success: {classification} (confidence: {confidence:.2f})")
                    return classification_data
                    
                except json.JSONDecodeError as jde:
                    logging.error(f"[CLASSIFY] {pdf_path_obj.name} - JSON decode failed: {str(jde)}")
                    return None
                except Exception as exc:
                    if attempt < _MAX_RETRIES - 1:
                        delay = _RETRY_DELAY_BASE_SECONDS * (2 ** attempt)
                        logging.warning(f"[CLASSIFY] {pdf_path_obj.name} - Error: {str(exc)[:100]}. Retrying in {delay}s...")
                        await asyncio.sleep(delay)
                        continue
                    logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Exhausted retries: {str(exc)[:150]}")
                    return None
                    
    except Exception as e:
        logging.error(f"[CLASSIFY] {pdf_path_obj.name} - Preprocessing error: {str(e)[:150]}")
        return None

async def extract_document_data_async(
    pdf_path: str,
    document_type: str,
    client: "genai.Client",
    semaphore: asyncio.Semaphore,
    json_responses_folder: str
) -> Optional[Dict[str, Any]]:
    """
    Extract data from a classified document using enhanced extraction prompts.
    
    Args:
        pdf_path: Path to the PDF file
        document_type: Classification result (employee_t&e or vendor_invoice)
        client: Shared genai client
        semaphore: Semaphore to control concurrent API calls
        json_responses_folder: Folder to save JSON responses
        
    Returns:
        Dictionary containing extracted data or None if failed
    """
    pdf_path_obj = Path(pdf_path)
    
    # Check page count - only process files with ≤20 pages (or first 20 pages if >20)
    total_pages = get_pdf_page_count(pdf_path)
    process_full_file = total_pages <= MAX_EXTRACTION_PAGES
    
    if total_pages > MAX_EXTRACTION_PAGES:
        logging.info(f"[EXTRACT] {pdf_path_obj.name} - Will use first {MAX_EXTRACTION_PAGES} pages: {total_pages} pages > {MAX_EXTRACTION_PAGES} page limit")
    
    # Select the appropriate extraction prompt
    if document_type == "employee_t&e":
        extraction_prompt = ENHANCED_EMPLOYEE_REIMBURSEMENT_EXTRACTION_PROMPT_V2
    elif document_type == "vendor_invoice":
        extraction_prompt = ENHANCED_VENDOR_INVOICE_EXTRACTION_PROMPT_V2
    else:
        logging.error(f"[EXTRACT] {pdf_path_obj.name} - Invalid document type: {document_type}")
        return None
    
    try:
        # Use full PDF for extraction or first 20 pages if file is oversized
        if process_full_file:
            pdf_bytes = pdf_path_obj.read_bytes()
        else:
            # Extract first 20 pages for oversized files
            pdf_bytes = extract_first_n_pages_pdf(pdf_path, MAX_EXTRACTION_PAGES)
        
        contents = [
            types.Part.from_bytes(
                data=pdf_bytes,
                mime_type='application/pdf',
            ),
            f"Extract detailed financial data from this {document_type.replace('_', ' ')} document with enhanced accuracy and completeness."
        ]
        
        config = types.GenerateContentConfig(system_instruction=extraction_prompt)
        
        async with semaphore:
            for attempt in range(_MAX_RETRIES):
                try:
                    logging.info(f"[EXTRACT] {pdf_path_obj.name} - Attempt {attempt + 1}/{_MAX_RETRIES} ({document_type}, {total_pages} pages)")
                    
                    response = await client.aio.models.generate_content(
                        model=EXTRACTION_MODEL,
                        contents=contents,
                        config=config
                    )
                    
                    # Save response for debugging in appropriate subfolder
                    extraction_log_filename = f"{pdf_path_obj.stem}_extraction_{document_type}_attempt_{attempt+1}.txt"
                    extraction_subfolder = os.path.join(json_responses_folder, document_type)
                    os.makedirs(extraction_subfolder, exist_ok=True)
                    extraction_log_path = os.path.join(extraction_subfolder, extraction_log_filename)
                    with open(extraction_log_path, 'w', encoding='utf-8') as f:
                        f.write(response.text)
                    
                    # Parse JSON response with robust error handling
                    resp_txt = response.text
                    json_start = resp_txt.find('{')
                    json_end = resp_txt.rfind('}') + 1
                    
                    if json_start == -1 or json_end == 0:
                        if any(err in resp_txt for err in ["502", "Service Unavailable", "server error"]):
                            raise RuntimeError("Transient upstream error – will retry")
                        logging.error(f"[EXTRACT] {pdf_path_obj.name} - Unparsable response: {resp_txt[:120]}")
                        return None
                    
                    try:
                        extraction_data = json.loads(resp_txt[json_start:json_end])
                    except json.JSONDecodeError as json_error:
                        # Try to fix common JSON malformations
                        logging.warning(f"[EXTRACT] {pdf_path_obj.name} - JSON malformed, attempting repair: {str(json_error)}")
                        try:
                            import re
                            json_str = resp_txt[json_start:json_end]
                            # Fix common issue: "text" (extra text) -> "text"
                            fixed_json = re.sub(r'"([^"]*)" \([^)]*\)', r'"\1"', json_str)
                            extraction_data = json.loads(fixed_json)
                            logging.info(f"[EXTRACT] {pdf_path_obj.name} - JSON repair successful")
                        except json.JSONDecodeError:
                            logging.error(f"[EXTRACT] {pdf_path_obj.name} - JSON repair failed")
                            raise json_error
                    
                    # Add metadata
                    extraction_data["file_name"] = pdf_path_obj.name
                    extraction_data["file_path"] = str(pdf_path_obj)
                    extraction_data["extraction_model"] = EXTRACTION_MODEL
                    extraction_data["document_type_processed"] = document_type
                    extraction_data["total_pages_in_pdf"] = total_pages
                    
                    logging.info(f"[EXTRACT] {pdf_path_obj.name} - Success ({document_type})")
                    return extraction_data
                    
                except json.JSONDecodeError as jde:
                    logging.error(f"[EXTRACT] {pdf_path_obj.name} - JSON decode failed: {str(jde)}")
                    return None
                except Exception as exc:
                    if attempt < _MAX_RETRIES - 1:
                        delay = _RETRY_DELAY_BASE_SECONDS * (2 ** attempt)
                        logging.warning(f"[EXTRACT] {pdf_path_obj.name} - Error: {str(exc)[:100]}. Retrying in {delay}s...")
                        await asyncio.sleep(delay)
                        continue
                    logging.error(f"[EXTRACT] {pdf_path_obj.name} - Exhausted retries: {str(exc)[:150]}")
                    return None
                    
    except Exception as e:
        logging.error(f"[EXTRACT] {pdf_path_obj.name} - Error: {str(e)[:150]}")
        return None

async def process_files_batch(
    files: List[Union[str, Tuple[str, str]]],
    client: "genai.Client",
    semaphore: asyncio.Semaphore,
    json_responses_folder: str,
    stage: str = "classification"
) -> Tuple[List[Dict[str, Any]], List[Dict[str, str]]]:
    """
    Unified function to process files through either classification or extraction stage.
    
    Args:
        files: List of file paths (for classification) or (file_path, doc_type) tuples (for extraction)
        client: Shared genai client
        semaphore: Semaphore to control concurrent API calls
        json_responses_folder: Folder to save JSON responses
        stage: "classification" or "extraction"
        
    Returns:
        Tuple of (successful_results, failed_files)
    """
    if not files:
        return [], []
    
    # Create appropriate async tasks based on stage
    if stage == "classification":
        tasks = [
            asyncio.create_task(classify_document_async(file_path, client, semaphore, json_responses_folder))
            for file_path in files
        ]
        task_items = files
        desc = f"Processing {len(files)} files - Classification"
    else:  # extraction
        tasks = [
            asyncio.create_task(extract_document_data_async(file_path, doc_type, client, semaphore, json_responses_folder))
            for file_path, doc_type in files
        ]
        task_items = files
        desc = f"Processing {len(files)} files - Extraction"
    
    successful_results = []
    failed_files = []
    
    with tqdm(total=len(tasks), desc=desc, unit="file") as pbar:
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        for item, result in zip(task_items, results):
            if stage == "classification":
                file_path = item
                filename = os.path.basename(file_path)
            else:  # extraction
                file_path, doc_type = item
                filename = os.path.basename(file_path)
            
            if isinstance(result, Exception):
                failed_files.append({
                    "file_name": filename,
                    "file_path": file_path,
                    "failure_stage": stage,
                    "error_message": str(result)[:200],
                    **({"doc_type": doc_type} if stage == "extraction" else {})
                })
                pbar.set_postfix_str(f"💥 {filename}")
                logging.error(f"[{stage.upper()} ERROR] {filename}: {str(result)[:100]}")
            elif result:
                # Check for preprocessing failures in classification
                if stage == "classification" and result.get("preprocessing_failure", False):
                    successful_results.append(result)
                    pbar.set_postfix_str(f"🔧 {filename} (preprocessing error)")
                else:
                    successful_results.append(result)
                    if stage == "classification":
                        classification = result.get("classification", "unknown")
                        confidence = result.get("confidence", 0)
                        total_pages = result.get("total_pages_in_pdf", "?")
                        pbar.set_postfix_str(f"✅ {filename} ({classification}, {confidence:.2f}, {total_pages}p)")
                    else:  # extraction
                        pbar.set_postfix_str(f"✅ {filename} ({doc_type})")
            else:
                failed_files.append({
                    "file_name": filename,
                    "file_path": file_path,
                    "failure_stage": stage,
                    "error_message": f"{stage.title()} returned None (JSON decode or API error)",
                    **({"doc_type": doc_type} if stage == "extraction" else {})
                })
                pbar.set_postfix_str(f"❌ {filename}")
                logging.warning(f"[{stage.upper()} FAILED] {filename}")
            
            pbar.update(1)
    
    return successful_results, failed_files

async def process_with_retries(
    files: List[Union[str, Tuple[str, str]]],
    client: "genai.Client",
    semaphore: asyncio.Semaphore,
    json_responses_folder: str,
    stage: str = "classification",
    max_passes: int = 3
) -> Tuple[List[Dict[str, Any]], List[Dict[str, str]]]:
    """
    Process files with automatic retries using the unified pipeline.
    
    Args:
        files: List of file paths (for classification) or (file_path, doc_type) tuples (for extraction)
        client: Shared genai client
        semaphore: Semaphore to control concurrent API calls
        json_responses_folder: Folder to save JSON responses
        stage: "classification" or "extraction"
        max_passes: Maximum number of retry passes
        
    Returns:
        Tuple of (all_successful_results, final_failed_files)
    """
    all_successful = []
    remaining_files = files
    
    for pass_num in range(max_passes):
        if not remaining_files:
            break
            
        pass_desc = f"Pass {pass_num + 1}/{max_passes}" if pass_num > 0 else "Initial processing"
        if pass_num > 0:
            logging.info(f"🔄 {pass_desc}: Retrying {len(remaining_files)} failed {stage} files...")
        
        successful, failed = await process_files_batch(
            remaining_files, client, semaphore, json_responses_folder, stage
        )
        
        all_successful.extend(successful)
        
        # If no failures, we're done
        if not failed:
            logging.info(f"✅ All {stage} tasks completed successfully!")
            break
            
        # For classification retries, filter out preprocessing failures (they shouldn't be retried)
        if stage == "classification":
            retryable_failed = [
                failure for failure in failed 
                if "preprocessing" not in failure["error_message"].lower()
            ]
            preprocessing_failures = [
                failure for failure in failed 
                if "preprocessing" in failure["error_message"].lower()
            ]
            
            if preprocessing_failures:
                logging.info(f"🔧 Skipping {len(preprocessing_failures)} preprocessing failures (not retryable)")
            
            remaining_files = [failure["file_path"] for failure in retryable_failed]
        else:  # extraction retries
            remaining_files = [
                (failure["file_path"], failure.get("doc_type", "unknown"))
                for failure in failed
            ]
        
        # If no retryable failures, we're done
        if not remaining_files:
            logging.info(f"✅ No retryable {stage} failures remaining")
            break
            
        if pass_num == max_passes - 1:
            logging.warning(f"⚠️  Reached maximum retry passes ({max_passes}) for {stage}")
    
    # Calculate final failures
    final_failed = []
    if stage == "classification":
        # Include both retryable failures that couldn't be recovered and preprocessing failures
        retryable_failed = [
            {"file_name": os.path.basename(file_path), "file_path": file_path, 
             "failure_stage": stage, "error_message": f"Failed after {max_passes} retry passes"}
            for file_path in remaining_files
        ]
        final_failed = retryable_failed
        # Add back preprocessing failures
        if 'preprocessing_failures' in locals():
            final_failed.extend(preprocessing_failures)
    else:
        final_failed = failed
    
    return all_successful, final_failed


def save_vendor_extraction_results_to_csv(extraction_results: List[Dict[str, Any]], output_file: str):
    """Save vendor invoice extraction results to separate CSV."""
    if not extraction_results:
        logging.warning("No vendor extraction results to save.")
        return
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    fieldnames = [
        "File Name", "File Path", "Document Type", "Readable", "Contains Invoices",
        "Multiple Documents", "Orientation Issues", "Data Source", "Issuer", 
        "Consignor", "Consignee", "Vendor Name", "Original Vendor Name", 
        "Invoice Type", "PAN", "Registration Numbers", "Invoice Date", 
        "Document Number", "Invoice Number", "Description", "Basic Amount", 
        "Tax Amount", "Total Amount", "Currency Code", "Original Amount", 
        "Exchange Rate", "Amount Calculated", "Calculation Method", "Is Main Invoice",
        "Total Pages In PDF", "Page Numbers", "Processing Notes"
    ]
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for result in extraction_results:
                doc_status = result.get("document_status", {})
                extracted_data = result.get("extracted_data", [])
                
                if not extracted_data:
                    # Create row with document status only
                    row = {
                        "File Name": result.get("file_name", ""),
                        "File Path": result.get("file_path", ""),
                        "Document Type": result.get("document_type_processed", ""),
                        "Readable": doc_status.get("readable", ""),
                        "Contains Invoices": doc_status.get("contains_invoices", ""),
                        "Multiple Documents": doc_status.get("multiple_documents", ""),
                        "Orientation Issues": doc_status.get("orientation_issues", ""),
                        "Data Source": "",
                        "Issuer": "",
                        "Consignor": "",
                        "Consignee": "",
                        "Vendor Name": "",
                        "Original Vendor Name": "",
                        "Invoice Type": "",
                        "PAN": "",
                        "Registration Numbers": "",
                        "Invoice Date": "",
                        "Document Number": "",
                        "Invoice Number": "",
                        "Description": "",
                        "Basic Amount": "",
                        "Tax Amount": "",
                        "Total Amount": "",
                        "Currency Code": "",
                        "Original Amount": "",
                        "Exchange Rate": "",
                        "Amount Calculated": "",
                        "Calculation Method": "",
                        "Is Main Invoice": "",
                        "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
                        "Page Numbers": "",
                        "Processing Notes": result.get("processing_notes", "")
                    }
                    writer.writerow(row)
                else:
                    # Create row for each extracted data entry
                    for data_entry in extracted_data:
                        reg_numbers = data_entry.get("registration_numbers", [])
                        reg_numbers_str = ", ".join([f"{reg.get('type', '')}:{reg.get('value', '')}" for reg in reg_numbers])
                        
                        page_numbers = data_entry.get("page_numbers", [])
                        page_numbers_str = ", ".join(map(str, page_numbers)) if page_numbers else ""
                        
                        row = {
                            "File Name": result.get("file_name", ""),
                            "File Path": result.get("file_path", ""),
                            "Document Type": result.get("document_type_processed", ""),
                            "Readable": doc_status.get("readable", ""),
                            "Contains Invoices": doc_status.get("contains_invoices", ""),
                            "Multiple Documents": doc_status.get("multiple_documents", ""),
                            "Orientation Issues": doc_status.get("orientation_issues", ""),
                            "Data Source": data_entry.get("data_source", ""),
                            "Issuer": data_entry.get("issuer", ""),
                            "Consignor": data_entry.get("consignor", ""),
                            "Consignee": data_entry.get("consignee", ""),
                            "Vendor Name": data_entry.get("vendor_name", ""),
                            "Original Vendor Name": data_entry.get("original_vendor_name", ""),
                            "Invoice Type": data_entry.get("invoice_type", ""),
                            "PAN": data_entry.get("pan", ""),
                            "Registration Numbers": reg_numbers_str,
                            "Invoice Date": data_entry.get("invoice_date", ""),
                            "Document Number": data_entry.get("document_number", ""),
                            "Invoice Number": data_entry.get("invoice_number", ""),
                            "Description": data_entry.get("description", ""),
                            "Basic Amount": data_entry.get("basic_amount", ""),
                            "Tax Amount": data_entry.get("tax_amount", ""),
                            "Total Amount": data_entry.get("total_amount", ""),
                            "Currency Code": data_entry.get("currency_code", ""),
                            "Original Amount": data_entry.get("original_amount", ""),
                            "Exchange Rate": data_entry.get("exchange_rate", ""),
                            "Amount Calculated": data_entry.get("amount_calculated", ""),
                            "Calculation Method": data_entry.get("calculation_method", ""),
                            "Is Main Invoice": data_entry.get("is_main_invoice", ""),
                            "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
                            "Page Numbers": page_numbers_str,
                            "Processing Notes": result.get("processing_notes", "")
                        }
                        writer.writerow(row)
        
        logging.info(f"Vendor extraction results saved to {output_file}")
    except Exception as e:
        logging.error(f"Error saving vendor extraction results: {str(e)}")

def save_employee_extraction_results_to_csv(extraction_results: List[Dict[str, Any]], output_file: str):
    """Save employee T&E extraction results to separate CSV."""
    if not extraction_results:
        logging.warning("No employee extraction results to save.")
        return
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    fieldnames = [
        "File Name", "File Path", "Document Type", "Readable", "Contains Invoices",
        "Multiple Documents", "Orientation Issues", "Data Source", "Vendor Name",
        "Employee Code", "Department", "PAN", "Registration Numbers", 
        "Invoice Date", "Document Number", "Invoice Number", "Description", 
        "Basic Amount", "Tax Amount", "Total Amount", "Currency Code", 
        "Original Amount", "Amount Calculated", "Calculation Method", 
        "Total Pages In PDF", "Page Numbers", "Processing Notes"
    ]
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for result in extraction_results:
                doc_status = result.get("document_status", {})
                extracted_data = result.get("extracted_data", [])
                
                if not extracted_data:
                    # Create row with document status only
                    row = {
                        "File Name": result.get("file_name", ""),
                        "File Path": result.get("file_path", ""),
                        "Document Type": result.get("document_type_processed", ""),
                        "Readable": doc_status.get("readable", ""),
                        "Contains Invoices": doc_status.get("contains_invoices", ""),
                        "Multiple Documents": doc_status.get("multiple_documents", ""),
                        "Orientation Issues": doc_status.get("orientation_issues", ""),
                        "Data Source": "",
                        "Vendor Name": "",
                        "Employee Code": "",
                        "Department": "",
                        "PAN": "",
                        "Registration Numbers": "",
                        "Invoice Date": "",
                        "Document Number": "",
                        "Invoice Number": "",
                        "Description": "",
                        "Basic Amount": "",
                        "Tax Amount": "",
                        "Total Amount": "",
                        "Currency Code": "",
                        "Original Amount": "",
                        "Amount Calculated": "",
                        "Calculation Method": "",
                        "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
                        "Page Numbers": "",
                        "Processing Notes": result.get("processing_notes", "")
                    }
                    writer.writerow(row)
                else:
                    # Create row for each extracted data entry
                    for data_entry in extracted_data:
                        reg_numbers = data_entry.get("registration_numbers", [])
                        reg_numbers_str = ", ".join([f"{reg.get('type', '')}:{reg.get('value', '')}" for reg in reg_numbers])
                        
                        page_numbers = data_entry.get("page_numbers", [])
                        page_numbers_str = ", ".join(map(str, page_numbers)) if page_numbers else ""
                        
                        row = {
                            "File Name": result.get("file_name", ""),
                            "File Path": result.get("file_path", ""),
                            "Document Type": result.get("document_type_processed", ""),
                            "Readable": doc_status.get("readable", ""),
                            "Contains Invoices": doc_status.get("contains_invoices", ""),
                            "Multiple Documents": doc_status.get("multiple_documents", ""),
                            "Orientation Issues": doc_status.get("orientation_issues", ""),
                            "Data Source": data_entry.get("data_source", ""),
                            "Vendor Name": data_entry.get("vendor_name", ""),
                            "Employee Code": data_entry.get("employee_code", ""),
                            "Department": data_entry.get("department", ""),
                            "PAN": data_entry.get("pan", ""),
                            "Registration Numbers": reg_numbers_str,
                            "Invoice Date": data_entry.get("invoice_date", ""),
                            "Document Number": data_entry.get("document_number", ""),
                            "Invoice Number": data_entry.get("invoice_number", ""),
                            "Description": data_entry.get("description", ""),
                            "Basic Amount": data_entry.get("basic_amount", ""),
                            "Tax Amount": data_entry.get("tax_amount", ""),
                            "Total Amount": data_entry.get("total_amount", ""),
                            "Currency Code": data_entry.get("currency_code", ""),
                            "Original Amount": data_entry.get("original_amount", ""),
                            "Amount Calculated": data_entry.get("amount_calculated", ""),
                            "Calculation Method": data_entry.get("calculation_method", ""),
                            "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
                            "Page Numbers": page_numbers_str,
                            "Processing Notes": result.get("processing_notes", "")
                        }
                        writer.writerow(row)
        
        logging.info(f"Employee extraction results saved to {output_file}")
    except Exception as e:
        logging.error(f"Error saving employee extraction results: {str(e)}")


def save_results_to_csv(
    data: List[Dict[str, Any]], 
    output_file: str, 
    field_config: Dict[str, Any],
    data_name: str = "results"
):
    """
    Generic function to save results to CSV.
    
    Args:
        data: List of dictionaries containing the data
        output_file: Output CSV file path
        field_config: Configuration dict with 'fieldnames' and 'row_mapper' function
        data_name: Name for logging (e.g., "classification results", "failed files")
    """
    if not data:
        logging.info(f"No {data_name} to save.")
        return
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    fieldnames = field_config['fieldnames']
    row_mapper = field_config.get('row_mapper')
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            if row_mapper:
                # Check if first row is a dict or list to determine writer type
                first_row = row_mapper(data[0]) if data else {}
                
                if isinstance(first_row, dict):
                    # Use DictWriter for dictionary mappings
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for item in data:
                        rows = row_mapper(item)
                        if isinstance(rows, list):
                            for row in rows:
                                writer.writerow(row)
                        else:
                            writer.writerow(rows)
                else:
                    # Use regular writer for list mappings
                    writer = csv.writer(csvfile)
                    writer.writerow(fieldnames)
                    
                    for item in data:
                        row = row_mapper(item)
                        writer.writerow(row)
            else:
                # Use simple writer for direct field mapping
                writer = csv.writer(csvfile)
                writer.writerow(fieldnames)
                for item in data:
                    row = [item.get(field, "") for field in fieldnames]
                    writer.writerow(row)
        
        logging.info(f"{data_name.title()} saved to {output_file}")
    except Exception as e:
        logging.error(f"Error saving {data_name}: {str(e)}")

# CSV Configuration Objects
def get_classification_csv_config():
    """Configuration for classification results CSV."""
    fieldnames = [
        "File Name", "File Path", "Classification", "Confidence", "Reasoning",
        "Key Indicators", "Has Employee Codes", "Has Vendor Letterhead", 
        "Has Invoice Numbers", "Has Travel Dates", "Appears Financial",
        "Has Amount Calculations", "Has Tax Information", "Contains Multiple Doc Types",
        "Primary Document Type", "Total Pages In PDF", "Pages Analyzed", "Classification Notes"
    ]
    
    def row_mapper(result):
        doc_chars = result.get("document_characteristics", {})
        key_indicators = result.get("key_indicators", [])
        
        return {
            "File Name": result.get("file_name", ""),
            "File Path": result.get("file_path", ""),
            "Classification": result.get("classification", ""),
            "Confidence": result.get("confidence", ""),
            "Reasoning": result.get("reasoning", ""),
            "Key Indicators": ", ".join(key_indicators) if key_indicators else "",
            "Has Employee Codes": doc_chars.get("has_employee_codes", ""),
            "Has Vendor Letterhead": doc_chars.get("has_vendor_letterhead", ""),
            "Has Invoice Numbers": doc_chars.get("has_invoice_numbers", ""),
            "Has Travel Dates": doc_chars.get("has_travel_dates", ""),
            "Appears Financial": doc_chars.get("appears_financial", ""),
            "Has Amount Calculations": doc_chars.get("has_amount_calculations", ""),
            "Has Tax Information": doc_chars.get("has_tax_information", ""),
            "Contains Multiple Doc Types": doc_chars.get("contains_multiple_doc_types", ""),
            "Primary Document Type": doc_chars.get("primary_document_type", ""),
            "Total Pages In PDF": result.get("total_pages_in_pdf", ""),
            "Pages Analyzed": result.get("pages_analyzed", ""),
            "Classification Notes": result.get("classification_notes", "")
        }
    
    return {"fieldnames": fieldnames, "row_mapper": row_mapper}

def get_failed_files_csv_config():
    """Configuration for failed files CSV."""
    fieldnames = ["Failed File Name", "Failed File Path", "Failure Stage", "Error Message"]
    
    def row_mapper(failure):
        return [
            failure["file_name"],
            failure["file_path"],
            failure["failure_stage"],
            failure["error_message"]
        ]
    
    return {"fieldnames": fieldnames, "row_mapper": row_mapper}

def create_worksheet(workbook, sheet_name: str, headers: List[str], data_rows: List[List], header_color: str = "366092"):
    """
    Helper function to create a worksheet with headers and data.
    
    Args:
        workbook: openpyxl workbook object
        sheet_name: Name of the worksheet
        headers: List of header strings
        data_rows: List of lists, each containing row data
        header_color: Hex color for header background (default: blue)
    """
    ws = workbook.create_sheet(sheet_name)
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    
    # Add data rows
    for row_idx, row_data in enumerate(data_rows, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    return ws

def create_excel_report(
    classification_results: List[Dict[str, Any]], 
    vendor_results: List[Dict[str, Any]], 
    employee_results: List[Dict[str, Any]], 
    output_folder: str, 
    input_folder_name: str
):
    """Create Excel file with 3 worksheets: Classification (problems only), Employee T&E, and Vendor Invoices."""
    excel_filename = f"{input_folder_name}_data.xlsx"
    excel_path = os.path.join(output_folder, excel_filename)
    
    try:
        # Create workbook and worksheets
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create Classification worksheet - only include irrelevant/problem files
        problem_classifications = [result for result in classification_results 
                                 if result.get("classification") in ["irrelevant", "processing_failed"]]
        
        if problem_classifications:
            headers = ["File Name", "Classification", "Total Pages In PDF", "Classification Notes"]
            data_rows = [
                [
                    result.get("file_name", ""),
                    result.get("classification", ""),
                    result.get("total_pages_in_pdf", ""),
                    result.get("classification_notes", "")
                ]
                for result in problem_classifications
            ]
            create_worksheet(workbook, "Classification", headers, data_rows, "366092")
        
        # Create Employee T&E worksheet
        if employee_results:
            ws_employee = workbook.create_sheet("Employee_T&E")
            
            # Headers for employee T&E (same as CSV)
            headers = [
                "File Name", "File Path", "Document Type", "Readable", "Contains Invoices",
                "Multiple Documents", "Orientation Issues", "Data Source", "Vendor Name",
                "Employee Code", "Department", "PAN", "Registration Numbers", 
                "Invoice Date", "Document Number", "Invoice Number", "Description", 
                "Basic Amount", "Tax Amount", "Total Amount", "Currency Code", 
                "Original Amount", "Amount Calculated", "Calculation Method", 
                "Total Pages In PDF", "Page Numbers", "Processing Notes"
            ]
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws_employee.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            # Add data
            row_idx = 2
            for result in employee_results:
                doc_status = result.get("document_status", {})
                extracted_data = result.get("extracted_data", [])
                
                if not extracted_data:
                    # Create row with document status only
                    row_data = [
                        result.get("file_name", ""),
                        result.get("file_path", ""),
                        result.get("document_type_processed", ""),
                        doc_status.get("readable", ""),
                        doc_status.get("contains_invoices", ""),
                        doc_status.get("multiple_documents", ""),
                        doc_status.get("orientation_issues", ""),
                        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                        result.get("total_pages_in_pdf", ""),
                        "",
                        result.get("processing_notes", "")
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws_employee.cell(row=row_idx, column=col, value=value)
                    row_idx += 1
                else:
                    # Create row for each extracted data entry
                    for data_entry in extracted_data:
                        reg_numbers = data_entry.get("registration_numbers", [])
                        reg_numbers_str = ", ".join([f"{reg.get('type', '')}:{reg.get('value', '')}" for reg in reg_numbers])
                        
                        page_numbers = data_entry.get("page_numbers", [])
                        page_numbers_str = ", ".join(map(str, page_numbers)) if page_numbers else ""
                        
                        row_data = [
                            result.get("file_name", ""),
                            result.get("file_path", ""),
                            result.get("document_type_processed", ""),
                            doc_status.get("readable", ""),
                            doc_status.get("contains_invoices", ""),
                            doc_status.get("multiple_documents", ""),
                            doc_status.get("orientation_issues", ""),
                            data_entry.get("data_source", ""),
                            data_entry.get("vendor_name", ""),
                            data_entry.get("employee_code", ""),
                            data_entry.get("department", ""),
                            data_entry.get("pan", ""),
                            reg_numbers_str,
                            data_entry.get("invoice_date", ""),
                            data_entry.get("document_number", ""),
                            data_entry.get("invoice_number", ""),
                            data_entry.get("description", ""),
                            data_entry.get("basic_amount", ""),
                            data_entry.get("tax_amount", ""),
                            data_entry.get("total_amount", ""),
                            data_entry.get("currency_code", ""),
                            data_entry.get("original_amount", ""),
                            data_entry.get("amount_calculated", ""),
                            data_entry.get("calculation_method", ""),
                            result.get("total_pages_in_pdf", ""),
                            page_numbers_str,
                            result.get("processing_notes", "")
                        ]
                        
                        for col, value in enumerate(row_data, 1):
                            ws_employee.cell(row=row_idx, column=col, value=value)
                        row_idx += 1
        
        # Create Vendor Invoices worksheet
        if vendor_results:
            ws_vendor = workbook.create_sheet("Vendor_Invoices")
            
            # Headers for vendor invoices (same as CSV)
            headers = [
                "File Name", "File Path", "Document Type", "Readable", "Contains Invoices",
                "Multiple Documents", "Orientation Issues", "Data Source", "Issuer", 
                "Consignor", "Consignee", "Vendor Name", "Original Vendor Name", 
                "Invoice Type", "PAN", "Registration Numbers", "Invoice Date", 
                "Document Number", "Invoice Number", "Description", "Basic Amount", 
                "Tax Amount", "Total Amount", "Currency Code", "Original Amount", 
                "Exchange Rate", "Amount Calculated", "Calculation Method", "Is Main Invoice",
                "Total Pages In PDF", "Page Numbers", "Processing Notes"
            ]
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws_vendor.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D97230", end_color="D97230", fill_type="solid")
            
            # Add data
            row_idx = 2
            for result in vendor_results:
                doc_status = result.get("document_status", {})
                extracted_data = result.get("extracted_data", [])
                
                if not extracted_data:
                    # Create row with document status only
                    row_data = [
                        result.get("file_name", ""),
                        result.get("file_path", ""),
                        result.get("document_type_processed", ""),
                        doc_status.get("readable", ""),
                        doc_status.get("contains_invoices", ""),
                        doc_status.get("multiple_documents", ""),
                        doc_status.get("orientation_issues", ""),
                        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                        result.get("total_pages_in_pdf", ""),
                        "",
                        result.get("processing_notes", "")
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws_vendor.cell(row=row_idx, column=col, value=value)
                    row_idx += 1
                else:
                    # Create row for each extracted data entry
                    for data_entry in extracted_data:
                        reg_numbers = data_entry.get("registration_numbers", [])
                        reg_numbers_str = ", ".join([f"{reg.get('type', '')}:{reg.get('value', '')}" for reg in reg_numbers])
                        
                        page_numbers = data_entry.get("page_numbers", [])
                        page_numbers_str = ", ".join(map(str, page_numbers)) if page_numbers else ""
                        
                        row_data = [
                            result.get("file_name", ""),
                            result.get("file_path", ""),
                            result.get("document_type_processed", ""),
                            doc_status.get("readable", ""),
                            doc_status.get("contains_invoices", ""),
                            doc_status.get("multiple_documents", ""),
                            doc_status.get("orientation_issues", ""),
                            data_entry.get("data_source", ""),
                            data_entry.get("issuer", ""),
                            data_entry.get("consignor", ""),
                            data_entry.get("consignee", ""),
                            data_entry.get("vendor_name", ""),
                            data_entry.get("original_vendor_name", ""),
                            data_entry.get("invoice_type", ""),
                            data_entry.get("pan", ""),
                            reg_numbers_str,
                            data_entry.get("invoice_date", ""),
                            data_entry.get("document_number", ""),
                            data_entry.get("invoice_number", ""),
                            data_entry.get("description", ""),
                            data_entry.get("basic_amount", ""),
                            data_entry.get("tax_amount", ""),
                            data_entry.get("total_amount", ""),
                            data_entry.get("currency_code", ""),
                            data_entry.get("original_amount", ""),
                            data_entry.get("exchange_rate", ""),
                            data_entry.get("amount_calculated", ""),
                            data_entry.get("calculation_method", ""),
                            data_entry.get("is_main_invoice", ""),
                            result.get("total_pages_in_pdf", ""),
                            page_numbers_str,
                            result.get("processing_notes", "")
                        ]
                        
                        for col, value in enumerate(row_data, 1):
                            ws_vendor.cell(row=row_idx, column=col, value=value)
                        row_idx += 1
        
        # Ensure at least one worksheet exists
        if not workbook.worksheets:
            workbook.create_sheet("No_Data")
            workbook.active["A1"] = "No data available for Excel export"
        
        # Save workbook
        workbook.save(excel_path)
        logging.info(f"Excel report saved to {excel_path}")
        
    except Exception as e:
        logging.error(f"Error creating Excel report: {str(e)}")

def create_summary_report(
    classification_results: List[Dict[str, Any]], 
    vendor_results: List[Dict[str, Any]], 
    employee_results: List[Dict[str, Any]], 
    failed_files: List[Dict[str, str]], 
    output_folder: str, 
    input_folder_name: str,
    total_files: int,
    elapsed_time: float,
    relevant_documents: List[Tuple[str, str]]
):
    """Create markdown summary report of the processing results."""
    summary_filename = f"{input_folder_name}_summary.md"
    summary_path = os.path.join(output_folder, summary_filename)
    
    try:
        # Calculate statistics
        classified_files = len(classification_results)
        relevant_files = len(relevant_documents)
        extracted_files = len(vendor_results) + len(employee_results)
        vendor_extracted = len(vendor_results)
        employee_extracted = len(employee_results)
        
        # Classification distribution
        classification_counts = {}
        page_stats = {"≤10": 0, "11-20": 0, ">20": 0}
        
        for result in classification_results:
            classification = result.get("classification", "unknown")
            classification_counts[classification] = classification_counts.get(classification, 0) + 1
            
            total_pages = result.get("total_pages_in_pdf", 0)
            if total_pages <= 10:
                page_stats["≤10"] += 1
            elif total_pages <= 20:
                page_stats["11-20"] += 1
            else:
                page_stats[">20"] += 1
        
        # Count oversized files that were processed
        oversized_processed = sum(1 for result in classification_results 
                                 if result.get("classification") in ["employee_t&e", "vendor_invoice"] 
                                 and result.get("total_pages_in_pdf", 0) > MAX_EXTRACTION_PAGES)
        
        # Count failed files by stage
        classification_failures = sum(1 for f in failed_files if f["failure_stage"] == "classification")
        extraction_failures = sum(1 for f in failed_files if f["failure_stage"] == "extraction")
        
        # Generate markdown content
        markdown_content = f"""# Invoice Processing Summary Report

## Processing Overview
- **Input Folder**: `{input_folder_name}`
- **Processing Date**: {time.strftime('%Y-%m-%d %H:%M:%S')}
- **Total Processing Time**: {elapsed_time:.2f} seconds
- **Configuration**: Enhanced 2-Step Pipeline with {CLASSIFICATION_MODEL} + {EXTRACTION_MODEL}

## File Processing Statistics

### Overall Results
| Metric | Count | Percentage |
|--------|-------|------------|
| **Total Files Processed** | {total_files} | 100.0% |
| **Successfully Classified** | {classified_files} | {classified_files/total_files*100:.1f}% |
| **Eligible for Extraction** | {relevant_files} | {relevant_files/total_files*100:.1f}% |
| **Successfully Extracted** | {extracted_files} | {extracted_files/relevant_files*100:.1f}% of eligible |
| **Failed Files** | {len(failed_files)} | {len(failed_files)/total_files*100:.1f}% |

### Classification Distribution
"""
        
        for classification, count in classification_counts.items():
            percentage = (count / classified_files) * 100 if classified_files > 0 else 0
            markdown_content += f"- **{classification.replace('_', ' ').title()}**: {count} files ({percentage:.1f}%)\n"
        
        markdown_content += f"""
### Page Distribution
"""
        for page_range, count in page_stats.items():
            percentage = (count / classified_files) * 100 if classified_files > 0 else 0
            markdown_content += f"- **{page_range} pages**: {count} files ({percentage:.1f}%)\n"
        
        markdown_content += f"""
### Extraction Results
- **Vendor Invoices Extracted**: {vendor_extracted}
- **Employee T&E Extracted**: {employee_extracted}
- **Oversized Files Processed** (using first {MAX_EXTRACTION_PAGES} pages): {oversized_processed}

## Processing Details

### Pipeline Configuration
- **Classification Model**: {CLASSIFICATION_MODEL}
- **Extraction Model**: {EXTRACTION_MODEL}
- **Classification Pages**: First {MAX_CLASSIFICATION_PAGES} pages analyzed
- **Extraction Page Limit**: ≤{MAX_EXTRACTION_PAGES} pages (truncated if larger)
- **Max Concurrent API Calls**: {MAX_CONCURRENT_API_CALLS}

### Quality Metrics
- **Classification Success Rate**: {classified_files/total_files*100:.1f}%
- **Extraction Success Rate**: {extracted_files/relevant_files*100:.1f}% (of eligible files)
- **Overall Processing Success Rate**: {(total_files - len(failed_files))/total_files*100:.1f}%

"""
        
        if failed_files:
            markdown_content += f"""## Failed Files Analysis

### Failure Distribution
- **Classification Failures**: {classification_failures}
- **Extraction Failures**: {extraction_failures}

### Failed Files Details
"""
            for failure in failed_files:
                markdown_content += f"- **{failure['file_name']}**: {failure['failure_stage']} - {failure['error_message'][:100]}{'...' if len(failure['error_message']) > 100 else ''}\n"
        else:
            markdown_content += "## ✅ All Files Processed Successfully\nNo files failed during processing!\n"
        
        markdown_content += f"""
## Output Files Generated

### CSV Files
- `{CLASSIFICATION_CSV_FILE}` - Classification results for all files
- `{VENDOR_EXTRACTION_CSV_FILE}` - Vendor invoice extraction results
- `{EMPLOYEE_EXTRACTION_CSV_FILE}` - Employee T&E extraction results
"""
        
        if failed_files:
            markdown_content += f"- `{FAILED_CSV_FILE}` - Failed files list with error details\n"
        
        markdown_content += f"""
### Excel File
- `{input_folder_name}_data.xlsx` - Combined results with 3 worksheets:
  - Classification
  - Employee_T&E  
  - Vendor_Invoices

### Debug Information
- JSON responses saved to: `json_responses_2step_enhanced/{input_folder_name}/`
- Logs saved to: `{LOGS_FOLDER}/{LOG_FILE}`

---
*Report generated by Enhanced 2-Step Invoice Processing Pipeline*
"""
        
        # Write markdown file
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        
        logging.info(f"Summary report saved to {summary_path}")
        
    except Exception as e:
        logging.error(f"Error creating summary report: {str(e)}")

def create_dynamic_output_folders(input_folder: str, base_output_folder: str, base_json_folder: str) -> Tuple[str, str]:
    """Create dynamic output folders based on input folder name."""
    input_folder_name = os.path.basename(os.path.normpath(input_folder))
    
    # Create output folder with input folder name
    dynamic_output_folder = os.path.join(base_output_folder, input_folder_name)
    os.makedirs(dynamic_output_folder, exist_ok=True)
    
    # Create JSON responses folder with input folder name and subfolders for each extraction type
    dynamic_json_folder = os.path.join(base_json_folder, input_folder_name)
    os.makedirs(dynamic_json_folder, exist_ok=True)
    os.makedirs(os.path.join(dynamic_json_folder, "classification"), exist_ok=True)
    os.makedirs(os.path.join(dynamic_json_folder, "vendor_invoice"), exist_ok=True)
    os.makedirs(os.path.join(dynamic_json_folder, "employee_t&e"), exist_ok=True)
    
    return dynamic_output_folder, dynamic_json_folder

async def setup_processing_environment(
    input_folder: str, 
    base_output_folder: str, 
    logs_folder: str, 
    base_json_responses_folder: str
) -> Tuple[str, str, List[str], "genai.Client", asyncio.Semaphore]:
    """
    Set up the processing environment including folders, logging, file discovery, and client.
    
    Returns:
        Tuple of (output_folder, json_responses_folder, pdf_files, client, semaphore)
    """
    # Create dynamic folder structure
    output_folder, json_responses_folder = create_dynamic_output_folders(
        input_folder, base_output_folder, base_json_responses_folder
    )
    
    # Ensure base directories exist
    os.makedirs(base_output_folder, exist_ok=True)
    os.makedirs(logs_folder, exist_ok=True)
    os.makedirs(base_json_responses_folder, exist_ok=True)
    
    # Set up logging
    setup_logging(logs_folder)
    
    # Clear old response artifacts for this input folder
    if os.path.exists(json_responses_folder):
        try:
            shutil.rmtree(json_responses_folder)
            logging.info(f"Cleared {json_responses_folder} directory.")
        except Exception as e:
            logging.error(f"Unable to clear {json_responses_folder}: {str(e)}")
    
    # Recreate folders with all subfolders
    os.makedirs(json_responses_folder, exist_ok=True)
    os.makedirs(os.path.join(json_responses_folder, "classification"), exist_ok=True)
    os.makedirs(os.path.join(json_responses_folder, "vendor_invoice"), exist_ok=True)
    os.makedirs(os.path.join(json_responses_folder, "employee_t&e"), exist_ok=True)
    
    # Find PDF files (including subdirectories)
    pdf_files = []
    for root, dirs, files in os.walk(input_folder):
        # Skip excluded directories
        dirs[:] = [d for d in dirs if d not in ["dup", "__pycache__"]]
        for file in files:
            if file.lower().endswith('.pdf'):
                pdf_files.append(os.path.join(root, file))
    
    if not pdf_files:
        logging.warning(f"No PDF files found in '{input_folder}'. Exiting.")
        return output_folder, json_responses_folder, [], None, None
    
    logging.info(f"🔍 Found {len(pdf_files)} PDFs to process in enhanced 2-step pipeline")
    logging.info(f"⚙️  Configuration: Max concurrent API calls = {MAX_CONCURRENT_API_CALLS}")
    logging.info(f"📄 Classification: {CLASSIFICATION_MODEL} (first {MAX_CLASSIFICATION_PAGES} pages)")
    logging.info(f"🤖 Extraction: {EXTRACTION_MODEL} (≤{MAX_EXTRACTION_PAGES} pages, first {MAX_EXTRACTION_PAGES} pages for oversized)")
    logging.info(f"📁 Output folder: {output_folder}")
    logging.info(f"🗂️  JSON responses: {json_responses_folder}")
    
    # Initialize genai client
    if os.getenv("GOOGLE_GENAI_USE_VERTEXAI", "").lower() == "true":
        client = genai.Client()
        logging.info("🔐 Using Vertex AI with Application Default Credentials")
    else:
        client = genai.Client(api_key=API_KEY)
        logging.info("🔐 Using regular Gemini API with API key")
    
    semaphore = asyncio.Semaphore(MAX_CONCURRENT_API_CALLS)
    
    return output_folder, json_responses_folder, pdf_files, client, semaphore

async def main(input_folder=None, output_folder=None, logs_folder=None, json_responses_folder=None):
    """
    Enhanced main 2-step processing function with improved accuracy and organization.
    Step 1: Classify documents using gemini-2.5-flash (first 7 pages)
    Step 2: Extract data from relevant documents ≤20 pages using gemini-2.5-pro with enhanced prompts
    """
    # Track execution time for summary report
    start_time = time.time()
    
    # Use provided arguments or fall back to defaults
    input_folder = input_folder or INPUT_FOLDER
    base_output_folder = output_folder or OUTPUT_FOLDER
    logs_folder = logs_folder or LOGS_FOLDER
    base_json_responses_folder = json_responses_folder or JSON_RESPONSES_FOLDER
    
    # Set up processing environment
    output_folder, json_responses_folder, pdf_files, client, semaphore = await setup_processing_environment(
        input_folder, base_output_folder, logs_folder, base_json_responses_folder
    )
    
    # Check if setup was successful
    if not pdf_files or client is None:
        return
    
    # ===========================
    # STEP 1: ENHANCED CLASSIFICATION WITH RETRIES
    # ===========================
    
    logging.info(f"🏷️  Starting Step 1: Enhanced Document Classification ({len(pdf_files)} files)")
    
    classification_results, failed_classification = await process_with_retries(
        pdf_files, client, semaphore, json_responses_folder, 
        stage="classification", max_passes=3
    )
    
    # Save classification results
    classification_csv_path = os.path.join(output_folder, CLASSIFICATION_CSV_FILE)
    save_results_to_csv(
        classification_results, 
        classification_csv_path, 
        get_classification_csv_config(),
        "classification results"
    )
    
    # Enhanced classification summary
    classification_counts = {}
    page_stats = {"≤10": 0, "11-20": 0, ">20": 0}
    
    for result in classification_results:
        classification = result.get("classification", "unknown")
        classification_counts[classification] = classification_counts.get(classification, 0) + 1
        
        total_pages = result.get("total_pages_in_pdf", 0)
        if total_pages <= 10:
            page_stats["≤10"] += 1
        elif total_pages <= 20:
            page_stats["11-20"] += 1
        else:
            page_stats[">20"] += 1
    
    logging.info(f"📊 Enhanced Classification Results:")
    total_classified = len(classification_results)
    for classification, count in classification_counts.items():
        percentage = (count / total_classified) * 100 if total_classified > 0 else 0
        logging.info(f"   {classification}: {count} files ({percentage:.1f}%)")
    
    logging.info(f"📖 Page Distribution:")
    for page_range, count in page_stats.items():
        percentage = (count / total_classified) * 100 if total_classified > 0 else 0
        logging.info(f"   {page_range} pages: {count} files ({percentage:.1f}%)")
    
    # ===========================
    # STEP 2: ENHANCED EXTRACTION WITH FILTERING
    # ===========================
    
    # Filter for relevant documents (all relevant docs, use first 20 pages for oversized)
    relevant_documents = []
    skipped_irrelevant = 0
    skipped_processing_failed = 0
    oversized_count = 0
    
    for result in classification_results:
        classification = result.get("classification", "")
        total_pages = result.get("total_pages_in_pdf", 0)
        file_path = result["file_path"]
        
        if classification == "irrelevant":
            skipped_irrelevant += 1
            continue
        elif classification == "processing_failed":
            skipped_processing_failed += 1
            continue
        elif classification in ["employee_t&e", "vendor_invoice"]:
            relevant_documents.append((file_path, classification))
            if total_pages > MAX_EXTRACTION_PAGES:
                oversized_count += 1
                logging.info(f"[PROCESS] {os.path.basename(file_path)} - Will use first {MAX_EXTRACTION_PAGES} pages ({total_pages} total pages)")
    
    logging.info(f"💎 Step 2: Enhanced Extraction from {len(relevant_documents)} eligible documents")
    logging.info(f"⚡ Processing Information:")
    logging.info(f"   📋 Irrelevant documents skipped: {skipped_irrelevant}")
    logging.info(f"   💥 Processing failed documents skipped: {skipped_processing_failed}")
    logging.info(f"   📚 Oversized files (using first {MAX_EXTRACTION_PAGES} pages): {oversized_count}")
    logging.info(f"   💰 Total relevant documents processing: {len(relevant_documents)} files")
    
    if not relevant_documents:
        logging.info("No eligible documents found for extraction. Process complete.")
        return
    
    # ===========================
    # STEP 2: ENHANCED EXTRACTION WITH RETRIES  
    # ===========================
    
    extraction_results, failed_extraction = await process_with_retries(
        relevant_documents, client, semaphore, json_responses_folder,
        stage="extraction", max_passes=3
    )
    
    # Note: Retries are now handled automatically within process_with_retries()
    
    # ===========================
    # SAVE FINAL RESULTS TO CSV
    # ===========================
    
    # Separate results by document type for separate CSV files
    vendor_results = [r for r in extraction_results if r.get("document_type_processed") == "vendor_invoice"]
    employee_results = [r for r in extraction_results if r.get("document_type_processed") == "employee_t&e"]
    
    # Save separate extraction results
    vendor_csv_path = os.path.join(output_folder, VENDOR_EXTRACTION_CSV_FILE)
    employee_csv_path = os.path.join(output_folder, EMPLOYEE_EXTRACTION_CSV_FILE)
    
    save_vendor_extraction_results_to_csv(vendor_results, vendor_csv_path)
    save_employee_extraction_results_to_csv(employee_results, employee_csv_path)
    
    # ===========================
    # ENHANCED FINAL SUMMARY
    # ===========================
    
    total_files = len(pdf_files)
    classified_files = len(classification_results)
    relevant_files = len(relevant_documents)
    extracted_files = len(extraction_results)
    vendor_extracted = len(vendor_results)
    employee_extracted = len(employee_results)
    
    # Count oversized files that were processed
    oversized_processed = sum(1 for result in classification_results 
                             if result.get("classification") in ["employee_t&e", "vendor_invoice"] 
                             and result.get("total_pages_in_pdf", 0) > MAX_EXTRACTION_PAGES)
    
    logging.info(f"🎯 ENHANCED 2-STEP FINAL RESULTS:")
    logging.info(f"   📁 Total files processed: {total_files}")
    logging.info(f"   🏷️  Successfully classified: {classified_files} ({classified_files/total_files*100:.1f}%)")
    logging.info(f"   💎 Eligible for extraction: {relevant_files} ({relevant_files/total_files*100:.1f}%)")
    logging.info(f"   ✅ Successfully extracted: {extracted_files} ({extracted_files/relevant_files*100:.1f}% of eligible)")
    logging.info(f"   🏢 Vendor invoices extracted: {vendor_extracted}")
    logging.info(f"   👤 Employee T&E extracted: {employee_extracted}")
    logging.info(f"   ⚡ Total non-processable files skipped: {total_files - relevant_files} files (irrelevant + processing failed)")
    logging.info(f"   📚 Oversized files processed with first {MAX_EXTRACTION_PAGES} pages: {oversized_processed}")
    
    # Save failed files (combine both classification and extraction failures)
    all_failed_files = failed_classification + failed_extraction
    # Remove duplicates based on file_path
    seen_files = set()
    unique_failed_files = []
    for failure in all_failed_files:
        if failure["file_path"] not in seen_files:
            unique_failed_files.append(failure)
            seen_files.add(failure["file_path"])
    
    if unique_failed_files:
        failed_csv_path = os.path.join(output_folder, FAILED_CSV_FILE)
        save_results_to_csv(
            unique_failed_files, 
            failed_csv_path, 
            get_failed_files_csv_config(),
            "failed files"
        )
        logging.warning(f"📋 {len(unique_failed_files)} files failed processing. See {failed_csv_path}")
    else:
        logging.info("🎉 All eligible files processed successfully!")
    
    # ===========================
    # GENERATE EXCEL AND SUMMARY REPORTS
    # ===========================
    
    # Get input folder name for file naming
    input_folder_name = os.path.basename(os.path.normpath(input_folder))
    
    # Create Excel report with 3 worksheets
    logging.info("📊 Creating Excel report with 3 worksheets...")
    create_excel_report(
        classification_results, 
        vendor_results, 
        employee_results, 
        output_folder, 
        input_folder_name
    )
    
    # Create markdown summary report
    logging.info("📝 Creating summary report...")
    # Calculate elapsed time
    elapsed_time = time.time() - start_time
    create_summary_report(
        classification_results,
        vendor_results, 
        employee_results,
        unique_failed_files,
        output_folder,
        input_folder_name,
        total_files,
        elapsed_time,
        relevant_documents
    )
    
    logging.info(f"📊 Enhanced results saved to {output_folder}/")
    logging.info(f"   - Classification: {CLASSIFICATION_CSV_FILE}")
    logging.info(f"   - Vendor Extraction: {VENDOR_EXTRACTION_CSV_FILE}")
    logging.info(f"   - Employee Extraction: {EMPLOYEE_EXTRACTION_CSV_FILE}")
    logging.info(f"   - Excel Report: {input_folder_name}_data.xlsx")
    logging.info(f"   - Summary Report: {input_folder_name}_summary.md")

if __name__ == "__main__":
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Enhanced 2-Step PDF processing: Classification then Specialized Extraction')
    parser.add_argument('--input', default=INPUT_FOLDER,
                       help=f'Input folder containing PDF files (default: {INPUT_FOLDER})')
    parser.add_argument('--output', default=OUTPUT_FOLDER,
                       help=f'Output folder for results (default: {OUTPUT_FOLDER})')
    parser.add_argument('--logs', default=LOGS_FOLDER,
                       help=f'Logs folder (default: {LOGS_FOLDER})')
    parser.add_argument('--json-responses', default=JSON_RESPONSES_FOLDER,
                       help=f'JSON responses folder (default: {JSON_RESPONSES_FOLDER})')
    
    args = parser.parse_args()
    
    # Set up basic logging
    setup_logging(args.logs)
    
    # Log configuration
    use_vertex = os.getenv("GOOGLE_GENAI_USE_VERTEXAI", "").lower() == "true"
    if use_vertex:
        project = os.getenv("GOOGLE_CLOUD_PROJECT", "not-set")
        location = os.getenv("GOOGLE_CLOUD_LOCATION", "not-set")
        logging.info(f"🚀 Starting Enhanced 2-Step Pipeline with Vertex AI - Project: {project}, Location: {location}")
    else:
        logging.info("🚀 Starting Enhanced 2-Step Pipeline with regular Gemini API")
    
    logging.info(f"📂 Input folder: {args.input}")
    logging.info(f"📁 Base output folder: {args.output}")
    logging.info(f"📋 Logs folder: {args.logs}")
    logging.info(f"🗂️  Base JSON responses folder: {args.json_responses}")
    logging.info(f"🔧 Enhanced features: 7-page classification, ≤10-page extraction, separate CSVs, dynamic folders")
    
    start_time = time.time()
    asyncio.run(main(
        input_folder=args.input,
        output_folder=args.output,
        logs_folder=args.logs,
        json_responses_folder=args.json_responses
    ))
    elapsed_time = time.time() - start_time
    logging.info(f"⏱️  Total enhanced 2-step execution time: {elapsed_time:.2f} seconds")